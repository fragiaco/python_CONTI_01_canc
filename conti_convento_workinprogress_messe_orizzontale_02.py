from tkinter import *
from tkinter import ttk
from tkinter import messagebox

import sqlite3
import pandas as pd
import os, sys, subprocess

from openpyxl.styles import Font, Alignment
from openpyxl.styles import Side, Border

from openpyxl import styles, formatting
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

# Leggi il file xlsx e trasformalo in dataframe impostando i nomi colonna
from openpyxl.worksheet.worksheet import Worksheet

from xlsxwriter.utility import xl_rowcol_to_cell
from pandastable import Table, TableModel, config

######################################################################
######## FUNZIONE CONNESSIONE AL DATABASE 'database_conti' ###########
######################################################################
def connessione():
    conn = sqlite3.connect('database_messe_orizzontale')

    cur = conn.cursor()
    try:
        cur.execute('''CREATE TABLE TABLE_Messe(ID integer not null PRIMARY KEY ,
                                                Anno integer not null ,
                                                Mese TEXT not null ,
                                                Nome_Celebrante TEXT not null ,
                                                Ad_Mentem integer not null ,
                                                Binate integer,
                                                Binate_Concelebrate integer,
                                                Trinate integer,
                                                Suffragi_Comunitari integer,
                                                Suffragi_Personali integer,
                                                Devozione integer,
                                                Benefattori integer,
                                                Pro_Populo integer)''')



    except:
        pass

    print(conn)
    print('Sei connesso al database_conti')
    conn.commit()
    conn.close()


connessione()

############################
######## TKINTER ###########
############################

root = Tk()

# DEFINISCO le dimensioni della GUI e il TITOLO nella barra superiore
height = 950  # altezza
width = 1680  # larghezza
top = 0
left = (root.winfo_screenwidth() - width) / 2
geometry = ("{}x{}+{}+{}".format(width, height, int(left), int(top)))
root.geometry(geometry)
root.resizable(0, 0)
root.title('Registro Messe')

foreground_Bianco = '#ffeddb'
background_Blu = 'blue'
# Label title
title = Label(root, text='Registro Messe', font=('verdana', 40, 'bold'), bg=background_Blu, fg=foreground_Bianco)
title.pack(side=TOP, fill=X)

###################################
######## TKINTER frames ###########
###################################

# Frame Combo - Top
Frame_combo = Frame(root, bd='4', bg=background_Blu, relief=RIDGE)
Frame_combo.place(x=5, y=73, width=1670, height=60)

# Frame Treeview - treeview right Frame
Frame_tree = Frame(root, bd='4', bg=background_Blu, relief=RIDGE)
Frame_tree.place(x=5, y=132, width=1570, height=335)
Frame_tree_Buttons= Frame(root, bd='4', bg=background_Blu, relief=RIDGE)
Frame_tree_Buttons.place(x=1575, y=132, width=100, height=335)

# Frame Update - update right Frame
Frame_update = Frame(root, bd='4', bg=background_Blu, relief=RIDGE)
Frame_update.place(x=5, y=465, width=1670, height=60)

Frame_pandastable = Frame(root, bd='4', bg=background_Blu, relief=RIDGE)
Frame_pandastable.place(x=5, y=525, width=1210, height=220)



###############################################
######## TKINTER ENTRY Frame_combo ############
###############################################
Entry_Anno_combo_IntVar                 = IntVar()
Entry_Mese_combo_StringVar              = StringVar()
Entry_Nome_Celebrante_combo_StringVar   = StringVar()
Entry_Ad_Mentem_combo_IntVar            = IntVar()
Entry_Binate_combo_IntVar               = IntVar()
Entry_Binate_Conc_combo_IntVar          = IntVar()
Entry_Trinate_combo_IntVar              = IntVar()
Entry_Suffragi_Comunitari_combo_IntVar  = IntVar()
Entry_Suffragi_Personali_combo_IntVar   = IntVar()
Entry_Devozione_combo_IntVar            = IntVar()
Entry_Benefattori_combo_IntVar          = IntVar()
Entry_Pro_Populo_combo_IntVar           = IntVar()


Label_TOTALE_Numero_Messe= Label(Frame_combo, text=' ', font=('verdana', 8, 'bold'),
                                                bg=background_Blu, fg=foreground_Bianco)
Label_TOTALE_Numero_Messe.grid(row=2, column=12, columnspan=2, padx=40, pady=10)


def trace_when_Entry_widget_is_updated(self, *args):
    try:
        Label_TOTALE_Numero_Messe.config(text=' ', font=('verdana', 16, 'bold'), bg=background_Blu, fg=foreground_Bianco)
        value = Entry_Ad_Mentem_combo_IntVar.get()+\
                Entry_Binate_combo_IntVar.get()+\
                Entry_Binate_Conc_combo_IntVar.get()+\
                Entry_Trinate_combo_IntVar.get()+\
                Entry_Suffragi_Comunitari_combo_IntVar.get()+\
                Entry_Suffragi_Personali_combo_IntVar.get()+\
                Entry_Devozione_combo_IntVar.get()+\
                Entry_Benefattori_combo_IntVar.get()+\
                Entry_Pro_Populo_combo_IntVar.get()

        text = "{}".format(value) if value else " "
        Label_TOTALE_Numero_Messe.config(text=text)

    except:
        pass


Entry_Anno_combo_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated)
Entry_Mese_combo_StringVar.trace_variable('w', trace_when_Entry_widget_is_updated)
Entry_Nome_Celebrante_combo_StringVar.trace_variable('w', trace_when_Entry_widget_is_updated)
Entry_Ad_Mentem_combo_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated)
Entry_Binate_combo_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated)
Entry_Binate_Conc_combo_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated)
Entry_Trinate_combo_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated)
Entry_Suffragi_Comunitari_combo_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated)
Entry_Suffragi_Personali_combo_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated)
Entry_Devozione_combo_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated)
Entry_Benefattori_combo_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated)
Entry_Pro_Populo_combo_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated)

# List Anni
Anni = [2020,2021,2022,2023,2024,2025,2026,2027,2028,2029,2030]

# Dropbox Anno
Entry_Anno_combo = ttk.Combobox(Frame_combo, font=("Helvetica", 10), values=Anni, textvariable=Entry_Anno_combo_IntVar)
Entry_Anno_combo.current(4)
Entry_Anno_combo.grid(row=2, column=0)
Entry_Anno_combo['state'] = 'readonly'


# List Mesi
Mesi = ["gennaio",
        "febbraio",
        "marzo",
        "aprile",
        "maggio",
        "giugno",
        "luglio",
        "agosto",
        "settembre",
        "ottobre",
        "novembre",
        "dicembre",
        ]

Celebranti = ["fra Antonio Porfiri",
        "fra Giacomo Rotunno",
        "fra Gabriele Giobbi",
        "Fra Alberto Dos Santos",
        "Ospite"]

# Dropbox Mesi
Entry_Mese_combo = ttk.Combobox(Frame_combo, font=("Helvetica", 10), values=Mesi, textvariable=Entry_Mese_combo_StringVar)
Entry_Mese_combo.current(0)
Entry_Mese_combo.grid(row=2, column=1)
Entry_Mese_combo['state'] = 'readonly'


Entry_Nome_Celebrante_combo = ttk.Combobox(Frame_combo, font=("Helvetica", 10), values=Celebranti, textvariable=Entry_Nome_Celebrante_combo_StringVar)
Entry_Nome_Celebrante_combo.current(1)
Entry_Nome_Celebrante_combo.grid(row=2, column=2)


Entry_Ad_Mentem_combo = Spinbox(Frame_combo, from_=0, to=31, wrap=True, width=11, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Ad_Mentem_combo_IntVar)
Entry_Ad_Mentem_combo.grid\
    (row=2, column=3)
Entry_Binate_combo = Spinbox(Frame_combo,from_=0, to=31,wrap=True, width=10, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Binate_combo_IntVar)
Entry_Binate_combo.grid\
    (row=2, column=4)
Entry_Binate_Conc_combo = Spinbox(Frame_combo,from_=0, to=31,wrap=True, width=11, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Binate_Conc_combo_IntVar)
Entry_Binate_Conc_combo.grid\
    (row=2, column=5)
Entry_Trinate_combo = Spinbox(Frame_combo,from_=0, to=31,wrap=True, width=10, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Trinate_combo_IntVar)
Entry_Trinate_combo.grid\
    (row=2, column=6)
Entry_Suffragi_Comunitari_combo = Spinbox(Frame_combo, from_=0, to=31,wrap=True, width=11,font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Suffragi_Comunitari_combo_IntVar)
Entry_Suffragi_Comunitari_combo.grid\
    (row=2, column=7)
Entry_Suffragi_Personali_combo = Spinbox(Frame_combo, from_=0, to=31,wrap=True, width=10,font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Suffragi_Personali_combo_IntVar)
Entry_Suffragi_Personali_combo.grid\
    (row=2, column=8)
Entry_Devozione_combo = Spinbox(Frame_combo,from_=0, to=31,wrap=True, width=11, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Devozione_combo_IntVar)
Entry_Devozione_combo.grid\
    (row=2, column=9)
Entry_Benefattori_combo = Spinbox(Frame_combo, from_=0, to=31,wrap=True, width=11,font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Benefattori_combo_IntVar)
Entry_Benefattori_combo.grid\
    (row=2, column=10)
Entry_Pro_Populo_combo = Spinbox(Frame_combo,from_=0, to=31,wrap=True, width=11, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Pro_Populo_combo_IntVar)
Entry_Pro_Populo_combo.grid\
    (row=2, column=11)





###############################################
######## TKINTER UPDATE Frame_UPDATE ############
###############################################
Entry_Anno_combo_update_IntVar              = IntVar()
Entry_Mese_combo_update_StringVar           = StringVar()
Entry_Nome_Celebrante_combo_update_StringVar= StringVar()
Entry_Ad_Mentem_combo_update_IntVar         = IntVar()
Entry_Binate_combo_update_update_IntVar     = IntVar()
Entry_Binate_Conc_update_combo_IntVar       = IntVar()
Entry_Trinate_combo_update_IntVar           = IntVar()
Entry_Suffragi_Comunitari_combo_update_IntVar= IntVar()
Entry_Suffragi_Personali_combo_update_IntVar = IntVar()
Entry_Devozione_combo_update_IntVar         = IntVar()
Entry_Benefattori_combo_update_IntVar       = IntVar()
Entry_Pro_Populo_combo_update_IntVar        = IntVar()



Label_TOTALE_Numero_Messe_update= Label(Frame_update, text=' ', font=('verdana', 8, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_TOTALE_Numero_Messe_update.grid(row=2, column=13, columnspan=2, padx=40, pady=10)


def trace_when_Entry_widget_is_updated_UPDATE(self, *args):
    try:
        Label_TOTALE_Numero_Messe_update.config(text=' ', font=('verdana', 16, 'bold'), bg=background_Blu, fg=foreground_Bianco)
        value = Entry_Ad_Mentem_combo_update_IntVar.get()+\
                Entry_Binate_combo_update_update_IntVar.get()+\
                Entry_Binate_Conc_update_combo_IntVar.get()+\
                Entry_Trinate_combo_update_IntVar.get()+\
                Entry_Suffragi_Comunitari_combo_update_IntVar.get()+\
                Entry_Suffragi_Personali_combo_update_IntVar.get()+\
                Entry_Devozione_combo_update_IntVar.get()+\
                Entry_Benefattori_combo_update_IntVar.get()+\
                Entry_Pro_Populo_combo_update_IntVar.get()

        text = "{}".format(value) if value else " "
        Label_TOTALE_Numero_Messe_update.config(text=text)

    except:
        pass


Entry_Anno_combo_update_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated_UPDATE)
Entry_Mese_combo_update_StringVar.trace_variable('w', trace_when_Entry_widget_is_updated_UPDATE)
Entry_Nome_Celebrante_combo_update_StringVar.trace_variable('w', trace_when_Entry_widget_is_updated_UPDATE)
Entry_Ad_Mentem_combo_update_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated_UPDATE)
Entry_Binate_combo_update_update_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated_UPDATE)
Entry_Binate_Conc_update_combo_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated_UPDATE)
Entry_Trinate_combo_update_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated_UPDATE)
Entry_Suffragi_Comunitari_combo_update_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated_UPDATE)
Entry_Suffragi_Personali_combo_update_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated_UPDATE)
Entry_Devozione_combo_update_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated_UPDATE)
Entry_Benefattori_combo_update_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated_UPDATE)
Entry_Pro_Populo_combo_update_IntVar.trace_variable('w', trace_when_Entry_widget_is_updated_UPDATE)



Entry_Id_combo_update = Entry(Frame_update, text=' ', font=('verdana', 8, 'bold'), width=5,bg=background_Blu, fg=foreground_Bianco)
Entry_Id_combo_update.grid(row=2, column=0)

# List Anni
Anni = [2020,2021,2022,2023,2024,2025,2026,2027,2028,2029,2030]

# Dropbox Anno
Entry_Anno_combo_update = ttk.Combobox(Frame_update, font=("Helvetica", 10), values=Anni, textvariable=Entry_Anno_combo_update_IntVar)
Entry_Anno_combo_update.current(4)
Entry_Anno_combo_update.grid(row=2, column=1)
Entry_Anno_combo_update['state'] = 'readonly'


# List Mesi
Mesi = ["gennaio",
        "febbraio",
        "marzo",
        "aprile",
        "maggio",
        "giugno",
        "luglio",
        "agosto",
        "settembre",
        "ottobre",
        "novembre",
        "dicembre",
        ]

Celebranti = ["fra Antonio Porfiri",
        "fra Giacomo Rotunno",
        "fra Gabriele Giobbi",
        "Fra Alberto Dos Santos",
        "Ospite"]

# Dropbox Mesi
Entry_Mese_combo_update = ttk.Combobox(Frame_update, font=("Helvetica", 10), values=Mesi, textvariable=Entry_Mese_combo_update_StringVar)
Entry_Mese_combo_update.current(0)
Entry_Mese_combo_update.grid(row=2, column=2)
Entry_Mese_combo_update['state'] = 'readonly'


Entry_Nome_Celebrante_combo_update = ttk.Combobox(Frame_update, font=("Helvetica", 10), values=Celebranti, textvariable=Entry_Nome_Celebrante_combo_update_StringVar)
Entry_Nome_Celebrante_combo_update.current(1)
Entry_Nome_Celebrante_combo_update.grid(row=2, column=3)


Entry_Ad_Mentem_combo_update = Spinbox(Frame_update, from_=0, to=31, wrap=True, width=7, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Ad_Mentem_combo_update_IntVar)
Entry_Ad_Mentem_combo_update.grid\
    (row=2, column=4)
Entry_Binate_combo_update = Spinbox(Frame_update,from_=0, to=31,wrap=True, width=11, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Binate_combo_update_update_IntVar)
Entry_Binate_combo_update.grid\
    (row=2, column=5)
Entry_Binate_Conc_combo_update = Spinbox(Frame_update,from_=0, to=31,wrap=True, width=10, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Binate_Conc_update_combo_IntVar)
Entry_Binate_Conc_combo_update.grid\
    (row=2, column=6)
Entry_Trinate_combo_update = Spinbox(Frame_update,from_=0, to=31,wrap=True, width=11, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Trinate_combo_update_IntVar)
Entry_Trinate_combo_update.grid\
    (row=2, column=7)
Entry_Suffragi_Comunitari_combo_update = Spinbox(Frame_update, from_=0, to=31,wrap=True, width=10,font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Suffragi_Comunitari_combo_update_IntVar)
Entry_Suffragi_Comunitari_combo_update.grid\
    (row=2, column=8)
Entry_Suffragi_Personali_combo_update = Spinbox(Frame_update, from_=0, to=31,wrap=True, width=10,font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Suffragi_Personali_combo_update_IntVar)
Entry_Suffragi_Personali_combo_update.grid\
    (row=2, column=9)
Entry_Devozione_combo_update = Spinbox(Frame_update,from_=0, to=31,wrap=True, width=11, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Devozione_combo_update_IntVar)
Entry_Devozione_combo_update.grid\
    (row=2, column=10)
Entry_Benefattori_combo_update = Spinbox(Frame_update, from_=0, to=31,wrap=True, width=11,font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Benefattori_combo_update_IntVar)
Entry_Benefattori_combo_update.grid\
    (row=2, column=11)
Entry_Pro_Populo_combo_update = Spinbox(Frame_update,from_=0, to=31,wrap=True, width=10, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE, textvariable=Entry_Pro_Populo_combo_update_IntVar)
Entry_Pro_Populo_combo_update.grid\
    (row=2, column=12)
#######################################################




############################
####### TREEVIEW ###########
############################



# Add some style
style = ttk.Style()
# Pick a theme
style.theme_use("default")

# Configure our treeview colors
style.configure("Treeview",
                background="#D3D3D3",
                foreground="black",
                rowheight=30,
                fieldbackground="#D3D3D3",
                font=('Calibri', 12)
                )

# Headings
style.configure("Treeview.Heading",
                font=('Calibri', 12, 'bold')
                )

# Change selected color
style.map('Treeview',
          background=[('selected', 'blue')]
          )

# Treeview Scrollbar
tree_scroll = Scrollbar(Frame_tree)
tree_scroll.pack(side=RIGHT, fill=Y)

# Create Treeview
my_tree = ttk.Treeview(Frame_tree, yscrollcommand=tree_scroll.set, selectmode="extended")
# Pack to the screen
my_tree.pack()

# Configure the scrollbar
tree_scroll.config(command=my_tree.yview)

# Define Our Columns
my_tree['columns'] = ("ID", "Anno", "Mese", "Nome_Celebrante", "Ad_Mentem", "Binate", "Binate_Concelebrate", "Trinate", "Suffragi_Comunitari", "Suffragi_Personali", "Devozione", "Benefattori", "Pro_Populo")





# Formate Our Columns
my_tree.column("#0", width=0, stretch=NO)
my_tree.column("ID", anchor=W, width=70)
my_tree.column("Anno", anchor=W, width=120)
my_tree.column("Mese", anchor=W, width=120)
my_tree.column("Nome_Celebrante", anchor=W, width=165)
my_tree.column("Ad_Mentem", anchor=W, width=120)
my_tree.column("Binate", anchor=W, width=120)
my_tree.column("Binate_Concelebrate", anchor=W, width=120)
my_tree.column("Trinate", anchor=W, width=120)
my_tree.column("Suffragi_Comunitari", anchor=W, width=120)
my_tree.column("Suffragi_Personali", anchor=W, width=120)
my_tree.column("Devozione", anchor=W, width=120)
my_tree.column("Benefattori", anchor=W, width=120)
my_tree.column("Pro_Populo", anchor=W, width=120)


# Create Headings
my_tree.heading("#0", text="", anchor=W)
my_tree.heading("ID", text="Id", anchor=W)
my_tree.heading("Anno", text="Anno", anchor=W)
my_tree.heading("Mese", text="Mese", anchor=W)
my_tree.heading("Nome_Celebrante", text="Celebrante", anchor=W)
my_tree.heading("Ad_Mentem", text="Ad_Ment", anchor=W)
my_tree.heading("Binate", text="Binate", anchor=W)
my_tree.heading("Binate_Concelebrate", text="Bin_Concel", anchor=W)
my_tree.heading("Trinate", text="Trinate", anchor=W)
my_tree.heading("Suffragi_Comunitari", text="Suffr_Com", anchor=W)
my_tree.heading("Suffragi_Personali", text="Suffr_Pers", anchor=W)
my_tree.heading("Devozione", text="Devozione", anchor=W)
my_tree.heading("Benefattori", text="Benefattori", anchor=W)
my_tree.heading("Pro_Populo", text="Pro_Populo", anchor=W)
#

############################
######## SQLITE3 ###########
############################

# Insert into TABLE_Conti
def submit():
    conn = sqlite3.connect('database_messe_orizzontale')
    cur = conn.cursor()

    #dati presi dalla combo di inserimento (non update)
    dati = [(Entry_Anno_combo_IntVar.get(),
             Entry_Mese_combo_StringVar.get(),
             Entry_Nome_Celebrante_combo_StringVar.get(),
             Entry_Ad_Mentem_combo_IntVar.get(),
             Entry_Binate_combo_IntVar.get(),
             Entry_Binate_Conc_combo_IntVar.get(),
             Entry_Trinate_combo_IntVar.get(),
             Entry_Suffragi_Comunitari_combo_IntVar.get(),
             Entry_Suffragi_Personali_combo_IntVar.get(),
             Entry_Devozione_combo_IntVar.get(),
             Entry_Benefattori_combo_IntVar.get(),
             Entry_Pro_Populo_combo_IntVar.get()
             )]


    cur.executemany(
        'INSERT INTO TABLE_Messe (Anno, Mese, Nome_Celebrante, Ad_Mentem, Binate, Binate_Concelebrate, Trinate, Suffragi_Comunitari, Suffragi_Personali, Devozione, Benefattori, Pro_Populo) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)', dati)



    conn.commit()
    # Close our connection
    conn.close()

def query_database():
    # Clear the Treeview
    for record in my_tree.get_children():
        my_tree.delete(record)

    # Create a database or connect to one that exists
    conn = sqlite3.connect('database_messe_orizzontale')

    # Create a cursor instance
    c = conn.cursor()

    c.execute("SELECT * FROM TABLE_Messe")
    records = c.fetchall()

    # for record in records:
    #     print(record)
    # record[0] = id key

    # COLORI RIGHE pari e dispari
    count = 0
    # Create striped row tags
    my_tree.tag_configure('oddrow', background="white")
    my_tree.tag_configure('evenrow', background="lightblue")

    for record in records:
        if count % 2 == 0:
            my_tree.insert(parent='', index=0, iid=record[0], text='',
                           values=(record[0], record[1], record[2], record[3], record[4], record[5], record[6], record[7], record[8], record[9], record[10], record[11], record[12]),
                           tags=('evenrow'))
        else:
            my_tree.insert(parent='', index=0, iid=record[0], text='',
                           values=(record[0], record[1], record[2], record[3], record[4], record[5], record[6], record[7],record[8], record[9], record[10], record[11], record[12]),
                           tags = ('oddrow'))
        count += 1

    # Al termine del processo la prima riga risulta evidenziata
    child_id = my_tree.get_children()[0]  # la prima riga dall'alto del treeview
    my_tree.focus(child_id)  # evidenziata
    my_tree.selection_set(child_id)

    # Commit changes
    conn.commit()

    # Close our connection
    conn.close()

def sqlite3_to_excel():
    # Create a database or connect to one that exists
    conn = sqlite3.connect('database_messe_orizzontale')

    # Create a cursor instance
    c = conn.cursor()

    query = "SELECT * FROM TABLE_Messe"  # query to collect recors

    df = pd.read_sql(query, conn)  # create dataframe

    df.sort_values(by='ID', ascending=False).to_excel('database_messe_orizzontale.xlsx', index=False, sheet_name='Dati')

    # Commit changes
    conn.commit()



    # Close our connection
    conn.close()


    ###########################################################
    ################# Creo il Workbook con OPENPYXL############
    ###########################################################


    wb = Workbook()
    wb = load_workbook(filename="database_messe_orizzontale.xlsx")
    ws = wb.active  # Worksheet
    ws.set_printer_settings(Worksheet.PAPERSIZE_A4, Worksheet.ORIENTATION_LANDSCAPE)
    ws.row_dimensions[1].height = 75
    # openpyxl freeze first row
    ws.freeze_panes = 'A2'
    # openpyxl filter columns
    ws.auto_filter.ref = ws.dimensions

    ############ RED ################
    red = NamedStyle(name="red")
    red.font = Font(name='Calibri', size=10, color='a81a1a', bold=True)
    red.alignment = Alignment(horizontal="center", vertical="center")
    red.fill = PatternFill('solid', fgColor='d1d22e')
    red.alignment = Alignment(textRotation=45)
    wb.add_named_style(red)

    ############ BLACK ################
    black = NamedStyle(name="black")
    black.font = Font(name='Calibri', size=10, color='000000', bold=True)
    black.alignment = Alignment(horizontal="left", vertical="center")
    wb.add_named_style(black)

    for row in ws[2:ws.max_row]:  # skip the header
        print(row) #(<Cell 'gennaio'.A7>, <Cell 'gennaio'.B7>, <Cell 'gennaio'.C7>, <Cell 'gennaio'.D7>)
        for cell in row: # il quarto valore della tuple
            #print(cell)  # <Cell 'multiple'.D7>
            cell.style= 'black'




    ws['A1'].style = 'red'
    ws['B1'].style = 'red'
    ws['C1'].style = 'red'
    ws['D1'].style = 'red'
    ws['E1'].style = 'red'
    ws['F1'].style = 'red'
    ws['G1'].style = 'red'
    ws['H1'].style = 'red'
    ws['I1'].style = 'red'
    ws['J1'].style = 'red'
    ws['K1'].style = 'red'
    ws['L1'].style = 'red'
    ws['M1'].style = 'red'
    ws['N1'].style = 'red'


    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 8
    ws.column_dimensions['K'].width = 8
    ws.column_dimensions['L'].width = 8
    ws.column_dimensions['M'].width = 8
    ws.column_dimensions['N'].width = 8


    # ws = wb.create_sheet('Dati')
    wb.save("database_messe_orizzontale_styled.xlsx")

    if sys.platform == "win32":
        os.startfile('database_messe_orizzontale_styled.xlsx')
    else:
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener, 'database_messe_orizzontale_styled.xlsx'])

def select_record(e):
    # Clear entry boxes
    Entry_Id_combo_update.delete(0, END)
    Entry_Anno_combo_update.delete(0, END)
    Entry_Mese_combo_update.delete(0, END)
    Entry_Nome_Celebrante_combo_update.delete(0, END)
    Entry_Ad_Mentem_combo_update.delete(0, END)
    Entry_Binate_combo_update.delete(0, END)
    Entry_Binate_Conc_combo_update.delete(0, END)
    Entry_Trinate_combo_update.delete(0, END)
    Entry_Suffragi_Comunitari_combo_update.delete(0, END)
    Entry_Suffragi_Personali_combo_update.delete(0, END)
    Entry_Devozione_combo_update.delete(0, END)
    Entry_Benefattori_combo_update.delete(0, END)
    Entry_Pro_Populo_combo_update.delete(0, END)

    # Grab record Number
    selected = my_tree.focus()  # focus restituisce l'ID key
    # print(selected) #esempio 38
    # Grab record values
    values = my_tree.item(selected, 'values')
    print(values) #esempio ('16', '2023', 'febbraio', 'fra Giacomo', 'Suffragi personali', '2')

    # outpus to entry boxes
    Entry_Id_combo_update.insert(0, values[0])  # 0 penso significa all'inizio
    Entry_Anno_combo_update.insert(0, values[1])
    Entry_Mese_combo_update.insert(0, values[2])
    Entry_Nome_Celebrante_combo_update.insert(0, values[3])
    Entry_Ad_Mentem_combo_update.insert(0, values[4])
    Entry_Binate_combo_update.insert(0, values[5])
    Entry_Binate_Conc_combo_update.insert(0, values[6])  # 0 penso significa all'inizio
    Entry_Trinate_combo_update.insert(0, values[7])
    Entry_Suffragi_Comunitari_combo_update.insert(0, values[8])
    Entry_Suffragi_Personali_combo_update.insert(0, values[9])
    Entry_Devozione_combo_update.insert(0, values[10])
    Entry_Benefattori_combo_update.insert(0, values[11])
    Entry_Pro_Populo_combo_update.insert(0, values[12])




# Bind the treeview ogni volta che seleziono una riga parte la funzione select_record
my_tree.bind("<ButtonRelease-1>", select_record)


#######################
def remove_one():
    # x = my_tree.selection()[0] #restituisce l'Id key
    x = my_tree.focus()
    my_tree.delete(x)

    # Create a database or connect to one that exists
    conn = sqlite3.connect('database_messe_orizzontale')

    # Create a cursor instance
    c = conn.cursor()

    # Delete From Database
    c.execute("DELETE from TABLE_Messe WHERE oid =" + Entry_Id_combo_update.get())

    # Commit changes
    conn.commit()

    # Close our connection
    conn.close()

    # Add a little message box for fun
    messagebox.showinfo("Deleted", "Riga Cancellata!")


#######################
######
# Update record
def update_record():
    # Grab the record number
    #print('update')
    selected = my_tree.focus()
    #print(selected)
    # Update record

    #Assegno i valori al Treeview
    my_tree.item(selected, text="", values=(
    Entry_Id_combo_update.get(),
    Entry_Anno_combo_update.get(),
    Entry_Mese_combo_update.get(),
    Entry_Nome_Celebrante_combo_update.get(),
    Entry_Ad_Mentem_combo_update.get(),
    Entry_Binate_combo_update.get(),
    Entry_Binate_Conc_combo_update.get(),
    Entry_Trinate_combo_update.get(),
    Entry_Suffragi_Comunitari_combo_update.get(),
    Entry_Suffragi_Personali_combo_update.get(),
    Entry_Devozione_combo_update.get(),
    Entry_Benefattori_combo_update.get(),
    Entry_Pro_Populo_combo_update.get()
    ))

    # Update the database
    # Create a database or connect to one that exists
    conn = sqlite3.connect('database_messe_orizzontale')
    #
    # Create a cursor instance
    c = conn.cursor()
    print(conn)

    c.execute("""   UPDATE TABLE_Messe 
                    SET
                    Anno = :Anno,
                    Mese = :Mese,
                    Nome_Celebrante = :Nome_Celebrante,
                    Ad_Mentem = :Ad_Mentem,
                    Binate = :Binate,
                    Binate_Concelebrate = :Binate_Concelebrate,
                    Trinate = :Trinate,
                    Suffragi_Comunitari = :Suffragi_Comunitari,
                    Suffragi_Personali = :Suffragi_Personali,
                    Devozione = :Devozione,
                    Benefattori = :Benefattori,
                    Pro_Populo = :Pro_Populo


     		        WHERE oid =""" + Entry_Id_combo_update.get(),
              {
                    'Anno':             Entry_Anno_combo_update.get(),
                    'Mese':             Entry_Mese_combo_update.get(),
                    'Nome_Celebrante':  Entry_Nome_Celebrante_combo_update.get(),
    		        'Ad_Mentem':        Entry_Ad_Mentem_combo_update.get(),
    		        'Binate':           Entry_Binate_combo_update.get(),
                    'Binate_Concelebrate': Entry_Binate_Conc_combo_update.get(),
    		        'Trinate':          Entry_Trinate_combo_update.get(),
    		        'Suffragi_Comunitari': Entry_Suffragi_Comunitari_combo_update.get(),
    		        'Suffragi_Personali': Entry_Suffragi_Personali_combo_update.get(),
    		        'Devozione':        Entry_Devozione_combo_update.get(),
                    'Benefattori':      Entry_Benefattori_combo_update.get(),
                    'Pro_Populo':       Entry_Pro_Populo_combo_update.get()
              })

    #    Commit changes
    conn.commit()
    #
    #         # Close our connection
    conn.close()
    # Add a little message box for fun
    messagebox.showinfo("Updated!", "Riga aggiornata!")

    #         # Clear entry boxes
    Entry_Id_combo_update.delete(0, END)
    Entry_Anno_combo_update.delete(0, END)
    Entry_Mese_combo_update.delete(0, END)
    Entry_Nome_Celebrante_combo_update.delete(0, END)
    Entry_Ad_Mentem_combo_update.delete(0, END)
    Entry_Binate_combo_update.delete(0, END)
    Entry_Binate_Conc_combo_update.delete(0, END)
    Entry_Trinate_combo_update.delete(0, END)
    Entry_Suffragi_Comunitari_combo_update.delete(0, END)
    Entry_Suffragi_Personali_combo_update.delete(0, END)
    Entry_Devozione_combo_update.delete(0, END)
    Entry_Benefattori_combo_update.delete(0, END)
    Entry_Pro_Populo_combo_update.delete(0, END)


# Read sqlite query results into a pandas DataFrame
con = sqlite3.connect("database_messe_orizzontale")
df = pd.read_sql_query("SELECT * from TABLE_Messe", con)

# Al termine del processo la prima riga risulta evidenziata
# child_id = my_tree.get_children()[0]  # la prima riga dall'alto del treeview
# my_tree.focus(child_id)  # evidenziata
# my_tree.selection_set(child_id)

#Grab record Number
#  
#print(values) #esempio ('16', '2023', 'febbraio', 'fra Giacomo', 'Suffragi personali', '2')
#values_list = list(values)
#print(values_list)

def aggiorna_dataframe_String_Int_Var():
        # Read sqlite query results into a pandas DataFrame
        con = sqlite3.connect("database_messe_orizzontale")
        df = pd.read_sql_query("SELECT * from TABLE_Messe", con)
        df = df.loc[(df['Anno']==Entry_Anno_combo_update_IntVar.get()) &
                    (df['Mese']==Entry_Mese_combo_update_StringVar.get())
                   ]
        return df

        con.close()

# Bind the treeview ogni volta che seleziono una riga parte la funzione select_record
#my_tree.bind("<ButtonRelease-1>", aggiorna_dataframe_String_Int_Var)

# print(df.to_markdown())
# print(df.nunique())
# #IMPORTANTE
# v0, v1= df.shape
# v2 = df.size
# print(f'In the df there are {v0} rows and {v1} columns and {v2} elements')
# print(f'Infatti {v0} x {v1} = {v2}')
# print(f"Nella colonna ID ci sono {df['ID'].size} righe")
#
# print(f"Nella colonna mese compaiono {df['Mese'].unique()} come valori unici")
# Verify that result of SQL query is stored in the dataframe

#df.drop(columns=['ID'])

pt = Table(Frame_pandastable, dataframe=df, width=100, height=100)



# df = pt.model.df
# df.drop(0)
pt.textcolor = 'blue'
pt.redraw()

options = {'colheadercolor':'blue','floatprecision': 1}
config.apply_options(options, pt)
pt.show()
con.close()

B_add = Button(Frame_tree_Buttons, text='aggiungi', width=10, command=lambda: [submit(), query_database()]).pack(side=TOP, pady=20)
B_excel = Button(Frame_tree_Buttons, text='Filtro_excel', width=10, command=sqlite3_to_excel).pack(side=TOP, pady=20)
B_update = Button(Frame_tree_Buttons, text='aggiorna', width=10, command=update_record).pack(side=TOP, pady=20)
B_delete = Button(Frame_tree_Buttons, text='cancella', width=10, command=remove_one).pack(side=TOP, pady=20)


query_database()
root.mainloop()