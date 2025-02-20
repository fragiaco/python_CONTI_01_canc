from tkinter import *
from tkinter import ttk
from tkinter import messagebox

import sqlite3
import pandas as pd
import os, sys, subprocess

from openpyxl.styles import Font, Alignment
from openpyxl.styles import Side, Border

from openpyxl import styles, formatting
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

# Leggi il file xlsx e trasformalo in dataframe impostando i nomi colonna
from openpyxl.worksheet.worksheet import Worksheet

from xlsxwriter.utility import xl_rowcol_to_cell


######################################################################
######## FUNZIONE CONNESSIONE AL DATABASE 'database_conti' ###########
######################################################################
def connessione():
    conn = sqlite3.connect('database_messe')

    cur = conn.cursor()
    try:
        cur.execute('''CREATE TABLE TABLE_Messe(ID integer not null PRIMARY KEY ,
                                                Anno integer not null ,
                                                Mese TEXT not null ,
                                                Nome TEXT not null ,
                                                Categoria TEXT not null ,
                                                
                                                Numero integer not null )''')
    except:
        pass

    print(conn)
    print('Sei connesso al database_conti')
    conn.commit()
    conn.close()


connessione()

############################
######## TKINTER ###########
############################

root = Tk()

# DEFINISCO le dimensioni della GUI e il TITOLO nella barra superiore
height = 950  # altezza
width = 1680  # larghezza
top = 0
left = (root.winfo_screenwidth() - width) / 2
geometry = ("{}x{}+{}+{}".format(width, height, int(left), int(top)))
root.geometry(geometry)
root.resizable(0, 0)
root.title('Registro Messe')

foreground_Bianco = '#ffeddb'
background_Blu = 'blue'
# Label title
title = Label(root, text='Registro Messe', font=('verdana', 40, 'bold'), bg=background_Blu, fg=foreground_Bianco)
title.pack(side=TOP, fill=X)

###################################
######## TKINTER frames ###########
###################################

# Frame Combo - left side Frame
Frame_combo = Frame(root, bd='4', bg=background_Blu, relief=RIDGE)
Frame_combo.place(x=20, y=75, width=450, height=450)
# Frame IN Combo - left side Frame
Frame_button_in_combo = Frame(Frame_combo, bd='4', bg=background_Blu, relief=RIDGE)
Frame_button_in_combo.place(y=400, width=445, height=42)


# Frame Treeview - treeview right Frame
Frame_tree = Frame(root, bd='4', bg=background_Blu, relief=RIDGE)
Frame_tree.place(x=475, y=75, width=730, height=450)
# Frame IN Treeview - left side Frame
Frame_button_in_Treeview = Frame(Frame_tree, bd='4', bg=background_Blu, relief=RIDGE)
Frame_button_in_Treeview.place(y=400, width=704, height=42)

# Frame Update - update right Frame
Frame_update = Frame(root, bd='4', bg=background_Blu, relief=RIDGE)
Frame_update.place(x=1210, y=75, width=450, height=450)
# Frame IN Update - left side Frame
Frame_button_in_update = Frame(Frame_update, bd='4', bg=background_Blu, relief=RIDGE)
Frame_button_in_update.place(y=400, width=445, height=42)

# Frame_bottom_left
Frame_bottom_left = Frame(root, bd='4', bg=background_Blu, relief=RIDGE) #azzurro fiordaliso
Frame_bottom_left.place(x=20, y=528, width=450, height=415)

# Frame Tabella - Tabella bottom Frame
Frame_tabella = Frame(root, bd='4', bg=background_Blu, relief=RIDGE)
Frame_tabella.place(x=475, y=528, width=730, height=415)

# Frame_bottom_right
Frame_bottom_right = Frame(root, bd='4', bg=background_Blu, relief=RIDGE)
Frame_bottom_right.place(x=1210, y=528, width=450, height=415)




#######################################################


#Labels in Frame Combo_insert
Label_combo_title = Label(Frame_combo, text='Inserisci Dati:', font=('verdana', 15, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_combo_title.grid(row=0, columnspan=2, padx=20, pady=10, sticky='w')

Label_riga_vuota = Label(Frame_combo, text='', font=('verdana', 5, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_riga_vuota.grid(row=1, columnspan=2, padx=20, pady=10, sticky='w')

Label_combo_anno = Label(Frame_combo, text='Anno', font=('verdana', 12, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_combo_anno.grid(row=2, padx=20, pady=10, sticky='w')

Label_combo_mese = Label(Frame_combo, text='Mese', font=('verdana', 12, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_combo_mese.grid(row=4, padx=20, pady=25, sticky='w')

Label_combo_Nome_celebrante = Label(Frame_combo, text='Nome Celebrante', font=('verdana', 12, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_combo_Nome_celebrante.grid(row=6, padx=25, pady=20, sticky='w')

Label_combo_categoria = Label(Frame_combo, text='Categoria', font=('verdana', 12, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_combo_categoria.grid(row=8, padx=25, pady=20, sticky='w')

Label_numero_messe = Label(Frame_combo, text='Numero Messe', font=('verdana', 12, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_numero_messe.grid(row=10, padx=25, pady=20, sticky='w')

#Labels in Frame Combo_update
Label_Combo_title_update = Label(Frame_update, text='Correggi - Cancella Dati:', font=('verdana', 15, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_Combo_title_update.grid(row=0, columnspan=2, padx=20, pady=10, sticky='w')

Label_Combo_subtitle_update = Label(Frame_update, text='Selezionare prima una riga nella tabella a sinistra',
                               font=('verdana', 10, 'bold'), bg='blue', fg='white')
Label_Combo_subtitle_update.grid(row=1, columnspan=2, padx=20, pady=10, sticky='w')

Id_label = Label(Frame_update, text="Id", font=('verdana', 12, 'bold'), bg='blue', fg='white')
Id_label.grid(row=2, column=0, padx=20, pady=12, sticky='w')

Label_combo_anno_update = Label(Frame_update, text='Anno', font=('verdana', 12, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_combo_anno_update.grid(row=3, padx=20, pady=12, sticky='w')

Label_mese = Label(Frame_update, text='Mese', font=('verdana', 12, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_mese.grid(row=4, padx=20, pady=12, sticky='w')

Label_combo_Nome_celebrante_update = Label(Frame_update, text='Nome Celebrante', font=('verdana', 12, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_combo_Nome_celebrante_update.grid(row=5, padx=20, pady=12, sticky='w')

Label_combo_categoria_update = Label(Frame_update, text='Categoria', font=('verdana', 12, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_combo_categoria_update.grid(row=6, padx=20, pady=12, sticky='w')

Label_numero_messe_update = Label(Frame_update, text='Numero Messe', font=('verdana', 12, 'bold'), bg=background_Blu, fg=foreground_Bianco)
Label_numero_messe_update.grid(row=7, padx=20, pady=12, sticky='w')


############################
####### COMBOBOX ###########
############################

# List Anno _ se metti gli apici li legge come stringhe
Anni = [2022,
        2023,
        2024,
        2025,
        ]

# List Mesi
Mesi = ["gennaio",
        "febbraio",
        "marzo",
        "aprile",
        "maggio",
        "giugno",
        "luglio",
        "agosto",
        "settembre",
        "ottobre",
        "novembre",
        "dicembre",
        ]

# List Entrate_Uscite
Nomi = ["fra Giacomo",
        "fra Gabriele",
                  ]

# List Categorie
Categorie   = [     "Ad Mentem",
                     "Binate",
                    "Binate concelebrate"
                     "Trinate",
                     "Suffragi personali",
                     "Suffragi obbligatori",
                     "Devozione",
                     "Benefattori",
                    "pro Populo",
                     ]




#############################
#####Combo insert############


# Dropbox Anno
anno_combo = ttk.Combobox(Frame_combo, font=("Helvetica", 12), values=Anni)
anno_combo.current(0)
anno_combo.grid(row=2, column=1)
anno_combo['state'] = 'readonly'

# Dropbox Mesi
mesi_combo = ttk.Combobox(Frame_combo, font=("Helvetica", 12), values=Mesi)
mesi_combo.current(0)
mesi_combo.grid(row=4, column=1)
mesi_combo['state'] = 'readonly'

# Dropbox Nomi
nomi_combo = ttk.Combobox(Frame_combo, font=("Helvetica", 12), values=Nomi)
nomi_combo.current(0)
nomi_combo.grid(row=6, column=1)
#nomi_combo['state'] = 'readonly'

# Dropbox Categoria
categorie_combo = ttk.Combobox(Frame_combo, font=("Helvetica", 12), values=Categorie)
categorie_combo.current(0)
categorie_combo.grid(row=8, column=1)
categorie_combo['state'] = 'readonly'

# Entry Numero
numero = Entry(Frame_combo, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE)
numero.grid(row=10, column=1)

###############################
##### Combo update ############

# anno_stringvar  = StringVar()
# mese_stringvar  = StringVar()
# nome_stringvar  = StringVar()
# categoria_stringvar = StringVar()
# numero_stringvar = StringVar()

Id_entry_toUpdate = Entry(Frame_update, font=('verdana', 12, 'bold'), bg='blue', fg='white', width=17)
Id_entry_toUpdate.grid(row=2, column=1)


# Dropbox Anno
anno_combo_update = ttk.Combobox(Frame_update, font=("Helvetica", 12), values=Anni)
#anno_combo_update.current(0)
anno_combo_update.grid(row=3, column=1)
#anno_combo_update['state'] = 'readonly'

# Dropbox Mesi
mesi_combo_update = ttk.Combobox(Frame_update, font=("Helvetica", 12), values=Mesi)
#mesi_combo_update.current(0)
mesi_combo_update.grid(row=4, column=1)
#mesi_combo_update['state'] = 'readonly'

# Dropbox Nomi
nomi_combo_update = ttk.Combobox(Frame_update, font=("Helvetica", 12), values=Nomi)
#nomi_combo_update.current(0)
nomi_combo_update.grid(row=5, column=1)
#nomi_combo_update['state'] = 'readonly'

# Dropbox Categoria
categorie_combo_update = ttk.Combobox(Frame_update, font=("Helvetica", 12), values=Categorie)
#categorie_combo_update.current(0)
categorie_combo_update.grid(row=6, column=1)
#categorie_combo_update['state'] = 'readonly'

# Entry Numero
numero_entry_toUpdate = Entry(Frame_update, font=("Helvetica", 12, 'bold'), bd=5, relief=GROOVE)
numero_entry_toUpdate.grid(row=7, column=1)






############################
####### TREEVIEW ###########
############################



# Add some style
style = ttk.Style()
# Pick a theme
style.theme_use("default")

# Configure our treeview colors
style.configure("Treeview",
                background="#D3D3D3",
                foreground="black",
                rowheight=30,
                fieldbackground="#D3D3D3",
                font=('Calibri', 12)
                )

# Headings
style.configure("Treeview.Heading",
                font=('Calibri', 12, 'bold')
                )

# Change selected color
style.map('Treeview',
          background=[('selected', 'blue')]
          )

# Treeview Scrollbar
tree_scroll = Scrollbar(Frame_tree)
tree_scroll.pack(side=RIGHT, fill=Y)

# Create Treeview
my_tree = ttk.Treeview(Frame_tree, yscrollcommand=tree_scroll.set, selectmode="extended")
# Pack to the screen
my_tree.pack()

# Configure the scrollbar
tree_scroll.config(command=my_tree.yview)

# Define Our Columns
my_tree['columns'] = ("Id", "Anno", "Mese", "Nome", "Categoria", "Numero")





# Formate Our Columns
my_tree.column("#0", width=0, stretch=NO)
my_tree.column("Id", anchor=W, width=81)
my_tree.column("Anno", anchor=W, width=80)
my_tree.column("Mese", anchor=W, width=80)
my_tree.column("Nome", anchor=W, width=200)
my_tree.column("Categoria", anchor=W, width=181)
my_tree.column("Numero", anchor=W, width=81)

# Create Headings
my_tree.heading("#0", text="", anchor=W)
my_tree.heading("Id", text="Id", anchor=W)
my_tree.heading("Anno", text="Anno", anchor=W)
my_tree.heading("Mese", text="Mese", anchor=W)
my_tree.heading("Nome", text="Nome del Celebrante", anchor=W)
my_tree.heading("Categoria", text="Messe", anchor=W)
my_tree.heading("Numero", text="Numero", anchor=W)




############################
######## SQLITE3 ###########
############################

# Insert into TABLE_Conti
def submit():
    conn = sqlite3.connect('database_messe')
    cur = conn.cursor()

    #dati presi dalla combo di inserimento (non update)
    dati = [(anno_combo.get(), mesi_combo.get(), nomi_combo.get(), categorie_combo.get(), numero.get())]

    cur.executemany(
        'INSERT INTO TABLE_Messe (Anno, Mese, Nome, Categoria, Numero) VALUES(?, ?, ?, ? ,?)', dati)
    conn.commit()
    # Close our connection
    conn.close()


# query_database ed insert rows into TREEVIEW
def query_database():
    # Clear the Treeview
    for record in my_tree.get_children():
        my_tree.delete(record)

    # Create a database or connect to one that exists
    conn = sqlite3.connect('database_messe')

    # Create a cursor instance
    c = conn.cursor()

    c.execute("SELECT * FROM TABLE_Messe")
    records = c.fetchall()

    # for record in records:
    #     print(record)
    # record[0] = id key

    # COLORI RIGHE pari e dispari
    count = 0
    # Create striped row tags
    my_tree.tag_configure('oddrow', background="white")
    my_tree.tag_configure('evenrow', background="lightblue")

    for record in records:
        if count % 2 == 0:
            my_tree.insert(parent='', index=0, iid=record[0], text='',
                           values=(record[0], record[1], record[2], record[3], record[4], record[5]),
                           tags=('evenrow'))
        else:
            my_tree.insert(parent='', index=0, iid=record[0], text='',
                           values=(record[0], record[1], record[2], record[3], record[4], record[5]),
                           tags=('oddrow'))

        count += 1

    # Al termine del processo la prima riga risulta evidenziata
    child_id = my_tree.get_children()[0]  # la prima riga dall'alto del treeview
    my_tree.focus(child_id)  # evidenziata
    my_tree.selection_set(child_id)

    # Commit changes
    conn.commit()

    # Close our connection
    conn.close()


def sqlite3_to_excel():
    # Create a database or connect to one that exists
    conn = sqlite3.connect('database_messe')

    # Create a cursor instance
    c = conn.cursor()

    query = "SELECT * FROM TABLE_Messe"  # query to collect recors

    df = pd.read_sql(query, conn)  # create dataframe

    df.sort_values(by='ID', ascending=False).to_excel('database_messe.xlsx', index=False, sheet_name='Dati')

    # Commit changes
    conn.commit()



    # Close our connection
    conn.close()



###########################################################
################# Creo il Workbook con OPENPYXL############
###########################################################


    wb = Workbook()
    wb = load_workbook(filename="database_messe.xlsx")
    ws = wb.active  # Worksheet

    ws.row_dimensions[1].height = 40
    # openpyxl freeze first row
    ws.freeze_panes = 'A2'
    # openpyxl filter columns
    ws.auto_filter.ref = ws.dimensions

    ############ RED ################
    red = NamedStyle(name="red")
    red.font = Font(name='Calibri', size=10, color='a81a1a', bold=True)
    red.alignment = Alignment(horizontal="center", vertical="center")
    red.fill = PatternFill('solid', fgColor='d1d22e')
    wb.add_named_style(red)

    ############ BLACK ################
    black = NamedStyle(name="black")
    black.font = Font(name='Calibri', size=10, color='000000', bold=True)
    black.alignment = Alignment(horizontal="left", vertical="center")
    wb.add_named_style(black)

    for row in ws[2:ws.max_row]:  # skip the header
        print(row) #(<Cell 'gennaio'.A7>, <Cell 'gennaio'.B7>, <Cell 'gennaio'.C7>, <Cell 'gennaio'.D7>)
        for cell in row: # il quarto valore della tuple
            print(cell)  # <Cell 'multiple'.D7>
            cell.style= 'black'




    ws['A1'].style = 'red'
    ws['B1'].style = 'red'
    ws['C1'].style = 'red'
    ws['D1'].style = 'red'
    ws['E1'].style = 'red'
    ws['F1'].style = 'red'
    #ws['G1'].style = 'red'

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 8
    #ws.column_dimensions['G'].width = 10



    # ws = wb.create_sheet('Dati')
    wb.save("database_messe_styled.xlsx")

    if sys.platform == "win32":
        os.startfile('database_messe_styled.xlsx')
    else:
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener, 'database_messe_styled.xlsx'])

'''

'''

'''
###########################################################
#################  EXCEL REPORT  ##########################
###########################################################


### Titolo 'Visualizza dati:'
Frame_excell_title = Label(Frame_excell, text='Visualizza dati:',
                           font=('verdana', 20, 'bold'), bg='blue', fg='#ffff66')
Frame_excell_title.grid(row=0, columnspan=2, padx=10, pady=10, sticky='w')
### Sottotitolo Scrivere prima
Frame_excell_subtitle_anno = Label(Frame_excell, text="Scrivere prima:",
                                   font=('verdana', 10, 'bold'), bg='blue', fg='white')
Frame_excell_subtitle_anno.grid(row=1, columnspan=2, padx=10, pady=10, sticky='w')
### Riga vuota
Frame_excell_subtitle_anno = Label(Frame_excell, text="",
                                   font=('verdana', 1, 'bold'), bg='blue', fg='white')
Frame_excell_subtitle_anno.grid(row=2, columnspan=2, padx=10, pady=10, sticky='w')
### Sottotitolo Anno
Frame_excell_subtitle_anno = Label(Frame_excell, text="Anno di interesse:",
                                   font=('verdana', 10, 'bold'), bg='blue', fg='white')
Frame_excell_subtitle_anno.grid(row=3, columnspan=2, padx=10, pady=10, sticky='w')
### Label Anno
Anno_label_excell = Label(Frame_excell, text="(esempio: 2022)", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Anno_label_excell.grid(row=4, column=0, padx=10, pady=10, sticky='w')
### Entry Anno
# Anno_entry_excell = Entry(Frame_excell, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable='')
# Anno_entry_excell.grid(row=4, column=1)
### Sottotitolo Saldo
Frame_excell_subtitle_anno_saldo = Label(Frame_excell, text="Saldo ad inizio anno:",
                                         font=('verdana', 10, 'bold'), bg='blue', fg='white')
Frame_excell_subtitle_anno_saldo.grid(row=5, columnspan=2, padx=10, pady=10, sticky='w')
### Label Saldo
Saldo_label_excell = Label(Frame_excell, text="(esempio: 99.99 oppure -99.99)", font=('verdana', 10, 'bold'), bg='blue',
                           fg='white')
Saldo_label_excell.grid(row=6, column=0, padx=10, pady=10, sticky='w')
### Entry Saldo
# Saldo_entry_excell = Entry(Frame_excell, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable='')
# Saldo_entry_excell.grid(row=6, column=1)

### Sottotitolo Residuo Messe
# Frame_excell_subtitle_residuo_messe = Label(Frame_excell, text="Dis/Avanzo Messe ad inizio anno",
#                             font=('verdana', 10, 'bold'), bg='blue', fg='white')
# Frame_excell_subtitle_residuo_messe.grid(row=7, columnspan=2, padx=10, pady=10, sticky='w')
# ### Label Residuo Messe
# Messe_residuo_label_excell = Label(Frame_excell, text="(esempio: 99 oppure -99)", font=('verdana', 10, 'bold'), bg='blue', fg='white')
# Messe_residuo_label_excell.grid(row=8, column=0, padx=10, pady=10, sticky='w')
# ### Entry Residuo Messe
# Messe_residuo_excell = Entry(Frame_excell, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable='')
# Messe_residuo_excell.grid(row=8, column=1)

'''

###########################################################
################# COMBOBOX UPDATE #########################
###########################################################




'''
Frame2_bottom_title = Label(Frame_combobox_ok, text='Correggi o Cancella:',
                            font=('verdana', 20, 'bold'), bg='blue', fg='#ffff66')
Frame2_bottom_title.grid(row=0, columnspan=2, padx=10, pady=10, sticky='w')

Frame2_bottom_subtitle = Label(Frame_combobox_ok, text='Selezionare prima una riga nella tabella sopra',
                               font=('verdana', 10, 'bold'), bg='blue', fg='white')
Frame2_bottom_subtitle.grid(row=1, columnspan=2, padx=10, pady=10, sticky='w')

Id_label = Label(Frame_combobox_ok, text="Id", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Id_label.grid(row=2, column=0, padx=10, pady=10, sticky='w')
Id_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white', width=17)
Id_entry.grid(row=2, column=1)

Anno_label = Label(Frame_combobox_ok, text="Anno", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Anno_label.grid(row=3, column=0, padx=10, pady=10, sticky='w')
Anno_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable=anno_stringvar)
# Anno_entry.grid(row=2, column=1, padx=10, pady=10)

Mese_label = Label(Frame_combobox_ok, text="Mese", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Mese_label.grid(row=4, column=0, padx=10, pady=10, sticky='w')
Mese_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable=mese_stringvar)
# Mese_entry.grid(row=3, column=1, padx=10, pady=10)

Entrate_Uscite_label = Label(Frame_combobox_ok, text="Entrate_Uscite", font=('verdana', 10, 'bold'), bg='blue',
                             fg='white')
Entrate_Uscite_label.grid(row=5, column=0, padx=10, pady=10, sticky='w')
Entrate_Uscite_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white',
                             textvariable=entrate_uscite_stringvar)
# Entrate_Uscite_entry.grid(row=4, column=1, padx=10, pady=10)
#
Categoria_label = Label(Frame_combobox_ok, text="Categoria", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Categoria_label.grid(row=6, column=0, padx=10, pady=10, sticky='w')
Categorie_Entrate_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white',
                                textvariable=categoria_stringvar)
# Categorie_Entrate_entry.grid(row=5, column=1, padx=10, pady=10)
#
Voce_label = Label(Frame_combobox_ok, text="Voce", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Voce_label.grid(row=7, column=0, padx=10, pady=10, sticky='w')
Voce_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable=voce_stringvar)
# Voce_entry.grid(row=6, column=1, padx=10, pady=10)
#
Euro_label = Label(Frame_combobox_ok, text="Euro", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Euro_label.grid(row=8, column=0, padx=10, pady=10, sticky='w')
Euro_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable=euro_stringvar)


# Euro_entry.grid(row=7, column=1, padx=10, pady=10)
'''

def select_record(e):
    # Clear entry boxes
    Id_entry_toUpdate.delete(0, END)
    anno_combo_update.delete(0, END)
    mesi_combo_update.delete(0, END)
    nomi_combo_update.delete(0, END)
    categorie_combo_update.delete(0, END)
    numero_entry_toUpdate.delete(0, END)

    # Grab record Number
    selected = my_tree.focus()  # focus restituisce l'ID key
    # print(selected) #esempio 38
    # Grab record values
    values = my_tree.item(selected, 'values')
    print(values) #esempio ('16', '2023', 'febbraio', 'fra Giacomo', 'Suffragi personali', '2')

    # outpus to entry boxes
    Id_entry_toUpdate.insert(0, values[0])  # 0 penso significa all'inizio
    anno_combo_update.insert(0, values[1])
    mesi_combo_update.insert(0, values[2])
    nomi_combo_update.insert(0, values[3])
    categorie_combo_update.insert(0, values[4])
    numero_entry_toUpdate.insert(0, values[5])


    # print(Anno_entry.get())


# Bind the treeview ogni volta che seleziono una riga parte la funzione select_record
my_tree.bind("<ButtonRelease-1>", select_record)


#######################
def remove_one():
    # x = my_tree.selection()[0] #restituisce l'Id key
    x = my_tree.focus()
    my_tree.delete(x)

    # Create a database or connect to one that exists
    conn = sqlite3.connect('database_messe')

    # Create a cursor instance
    c = conn.cursor()

    # Delete From Database
    c.execute("DELETE from TABLE_Messe WHERE oid=" + Id_entry_toUpdate.get())

    # Commit changes
    conn.commit()

    # Close our connection
    conn.close()

    # Add a little message box for fun
    messagebox.showinfo("Deleted", "Riga Cancellata!")


#######################
######
# Update record
def update_record():
    # Grab the record number
    #print('update')
    selected = my_tree.focus()
    #print(selected)
    # Update record
    my_tree.item(selected, text="", values=(
    Id_entry_toUpdate.get(), anno_combo_update.get(), mesi_combo_update.get(), nomi_combo_update.get(), categorie_combo_update.get(),
    numero_entry_toUpdate.get()))

    # Update the database
    # Create a database or connect to one that exists
    conn = sqlite3.connect('database_messe')
    #
    # Create a cursor instance
    c = conn.cursor()

    #
    c.execute("""UPDATE TABLE_Messe SET
    		Anno = :Anno,
    		Mese = :Mese,
    		Nome = :Nome,
    		Categoria = :Categoria,
    		Numero = :Numero

     		WHERE oid = :oid""",
              {
                  'Anno': anno_combo_update.get(),
                  'Mese': mesi_combo_update.get(),
                  'Nome': nomi_combo_update.get(),
                  'Categoria': categorie_combo_update.get(),
                  'Numero': numero_entry_toUpdate.get(),
                  'oid': Id_entry_toUpdate.get()
              })
    #
    #    Commit changes
    conn.commit()
    #
    #         # Close our connection
    conn.close()
    # Add a little message box for fun
    messagebox.showinfo("Updated!", "Riga aggiornata!")

    #         # Clear entry boxes
    Id_entry_toUpdate.delete(0, END)
    anno_combo_update.delete(0, END)
    mesi_combo_update.delete(0, END)
    nomi_combo_update.delete(0, END)
    categorie_combo_update.delete(0, END)
    numero_entry_toUpdate.delete(0, END)

'''

####

# class Complex:
#     def __init__(self, realpart, imagpart):
#         self.r = realpart
#         self.i = imagpart
# x = Complex(3.0, -4.5)
# x.r, x.i

anno_report_IntVar = IntVar()
saldo_report_DoubleVar = DoubleVar()

Anno_entry_excell = Entry(Frame_excell, font=('verdana', 10, 'bold'), bg='blue', fg='white',
                          textvariable=anno_report_IntVar)
Anno_entry_excell.grid(row=4, column=1)

Saldo_entry_excell = Entry(Frame_excell, font=('verdana', 10, 'bold'), bg='blue', fg='white',
                           textvariable=saldo_report_DoubleVar)
Saldo_entry_excell.grid(row=6, column=1)


# global anno_report

#################################################################
#########################    Class    ###########################
#################################################################

class Report():
    # if __name__ == '__main__':

    def __int__(self, anno, saldo):
        self.anno_report = anno
        self.saldo_report= saldo


    def anno_report_func(self):

         try:
            self.anno_report = anno_report_IntVar.get()
            return int(self.anno_report)  # return: altrimenti restituisce None - int: altrimenti Type STRING
         except:
             anno_report_IntVar.set(0)
             messagebox.showwarning(title='Dati Mancanti o Errati', message="Scrivere l'anno di interesse")

    def saldo_report_func(self):

         try:
            self.saldo_report = saldo_report_DoubleVar.get()
            return self.saldo_report  # return: altrimenti restituisce None - int: altrimenti Type STRING
         except:
             saldo_report_DoubleVar.set(0.0)
             messagebox.showwarning(title='Dati Mancanti o Errati', message="Scrivere il Saldo iniziale")

    def report(self):
      try:

        import pandas as pd

        import numpy as np

        import sqlite3

        import openpyxl
        from openpyxl import Workbook
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Font, Alignment
        from openpyxl.styles import Side, Border
        from openpyxl.styles import PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.worksheet.worksheet import Worksheet

        import os, sys, subprocess

        # CONNESSIONE A SQLITE3
        conn = sqlite3.connect('database_conti')
        cur = conn.cursor()

        # MI ASSIOCURO DI ESSERE CONNESSO
        # print(conn)
        # print('Sei connesso al database_conti')

        # CREO DATAFRAME df_database_conti
        df_database_conti = pd.read_sql("select * from TABLE_Conti", conn)
        # print(df_database_conti.info(verbose=True))

        # COMMIT e CLOSE
        conn.commit()
        conn.close()

        # LIST TITOLI DATAFRAMES 12 MESI
        list_df_conti_mese_entrate = ['df_database_conti_entrate_gennaio',
                                      'df_database_conti_entrate_febbraio',
                                      'df_database_conti_entrate_marzo',
                                      'df_database_conti_aprile',
                                      'df_database_conti_maggio',
                                      'df_database_conti_giugno',
                                      'df_database_conti_luglio',
                                      'df_database_conti_agosto',
                                      'df_database_conti_settembre',
                                      'df_database_conti_ottobre',
                                      'df_database_conti_novembre',
                                      'df_database_conti_dicembre'
                                      ]

        list_df_conti_mese_uscite = ['df_database_conti_uscite_gennaio',
                                     'df_database_conti_uscite_febbraio',
                                     'df_database_conti_uscite_marzo',
                                     'df_database_conti_uscite_aprile',
                                     'df_database_conti_uscite_maggio',
                                     'df_database_conti_uscite_giugno',
                                     'df_database_conti_uscite_luglio',
                                     'df_database_conti_uscite_agosto',
                                     'df_database_conti_uscite_settembre',
                                     'df_database_conti_uscite_ottobre',
                                     'df_database_conti_uscite_novembre',
                                     'df_database_conti_uscite_dicembre'
                                     ]

        list_mese = ['gennaio',
                     'febbraio',
                     'marzo',
                     'aprile',
                     'maggio',
                     'giugno',
                     'luglio',
                     'agosto',
                     'settembre',
                     'ottobre',
                     'novembre',
                     'dicembre'
                     ]

        ########### imposto anno ##############
        # anno = Report.anno_report_func(anno_report_Stringvar)
        # B_report = Button(Frame_excell_botton, text='report', width=10, command= lambda: print(Report.anno_report_func(anno_report_Stringvar))).grid(row=0, column=2, padx=20, pady=15)

        i = 0
        for x in range(12):
            list_df_conti_mese_entrate[i] = df_database_conti.loc[
                (df_database_conti['Anno'] == Report.anno_report_func(anno_report_IntVar)) &
                (df_database_conti['Mese'] == list_mese[i]) &
                (df_database_conti['Entrate_Uscite'] == 'Entrate')]
            # print(list_df_conti_camerino_pivot_entrate[i].head())

            list_df_conti_mese_uscite[i] = df_database_conti.loc[
                (df_database_conti['Anno'] == Report.anno_report_func(anno_report_IntVar)) &
                (df_database_conti['Mese'] == list_mese[i]) &
                (df_database_conti['Entrate_Uscite'] == 'Uscite')]
            # print(list_df_conti_camerino_mese_uscite[i].head())

            i += 1

        i = 0

        for x in range(12):

            dataframe_empty_list = [(i, Report.anno_report_func(anno_report_IntVar), list_mese[i], 'Entrate', 'vuoto', 'vuoto', 0)]
            if list_df_conti_mese_entrate[i].empty:
                list_df_conti_mese_entrate[i] = pd.DataFrame \
                    (dataframe_empty_list,
                     columns=['index', 'Anno', 'Mese', 'Entrate_Uscite', 'Categoria', 'Voce', 'Euro'])

            dataframe_empty_list = [(i, Report.anno_report_func(anno_report_IntVar), list_mese[i], 'Uscite', 'vuoto', 'vuoto', 0)]
            if list_df_conti_mese_uscite[i].empty:
                list_df_conti_mese_uscite[i] = pd.DataFrame \
                    (dataframe_empty_list,
                     columns=['index', 'Anno', 'Mese', 'Entrate_Uscite', 'Categoria', 'Voce', 'Euro'])

            print(list_df_conti_mese_entrate[i].to_markdown())
            print('')
            print(list_df_conti_mese_uscite[i].to_markdown())
            print('')

            # print(list_df_conti_mese_uscite[i].info())
            i += 1

        list_df_conti_mese_entrate = [list_df_conti_mese_entrate[0],  # df_conti_camerino_pivot_entrate_gennaio
                                      list_df_conti_mese_entrate[1],
                                      list_df_conti_mese_entrate[2],
                                      list_df_conti_mese_entrate[3],
                                      list_df_conti_mese_entrate[4],
                                      list_df_conti_mese_entrate[5],
                                      list_df_conti_mese_entrate[6],
                                      list_df_conti_mese_entrate[7],
                                      list_df_conti_mese_entrate[8],
                                      list_df_conti_mese_entrate[9],
                                      list_df_conti_mese_entrate[10],
                                      list_df_conti_mese_entrate[11]
                                      ]

        list_df_conti_mese_uscite = [list_df_conti_mese_uscite[0],  # df_conti_camerino_pivot_uscite_gennaio
                                     list_df_conti_mese_uscite[1],
                                     list_df_conti_mese_uscite[2],
                                     list_df_conti_mese_uscite[3],
                                     list_df_conti_mese_uscite[4],
                                     list_df_conti_mese_uscite[5],
                                     list_df_conti_mese_uscite[6],
                                     list_df_conti_mese_uscite[7],
                                     list_df_conti_mese_uscite[8],
                                     list_df_conti_mese_uscite[9],
                                     list_df_conti_mese_uscite[10],
                                     list_df_conti_mese_uscite[11]
                                     ]

        list_df_conti_mese_entrate = [list_df_conti_mese_entrate[0],  # df_conti_camerino_pivot_entrate_gennaio
                                      list_df_conti_mese_entrate[1],
                                      list_df_conti_mese_entrate[2],
                                      list_df_conti_mese_entrate[3],
                                      list_df_conti_mese_entrate[4],
                                      list_df_conti_mese_entrate[5],
                                      list_df_conti_mese_entrate[6],
                                      list_df_conti_mese_entrate[7],
                                      list_df_conti_mese_entrate[8],
                                      list_df_conti_mese_entrate[9],
                                      list_df_conti_mese_entrate[10],
                                      list_df_conti_mese_entrate[11]
                                      ]

        list_df_conti_mese_uscite = [list_df_conti_mese_uscite[0],  # df_conti_camerino_pivot_uscite_gennaio
                                     list_df_conti_mese_uscite[1],
                                     list_df_conti_mese_uscite[2],
                                     list_df_conti_mese_uscite[3],
                                     list_df_conti_mese_uscite[4],
                                     list_df_conti_mese_uscite[5],
                                     list_df_conti_mese_uscite[6],
                                     list_df_conti_mese_uscite[7],
                                     list_df_conti_mese_uscite[8],
                                     list_df_conti_mese_uscite[9],
                                     list_df_conti_mese_uscite[10],
                                     list_df_conti_mese_uscite[11]
                                     ]

        # print(list_df_conti_mese_entrate[0].columns)

        def pivot_table_w_subtotals(df, values, indices, columns, aggfunc, fill_value):
            listOfTable = []
            for indexNumber in range(len(indices)):
                n = indexNumber + 1
                if n == 1:
                    table = pd.pivot_table(df, values=values, index=indices[:n], columns=columns, aggfunc=aggfunc,
                                           fill_value=fill_value)

                else:
                    table = pd.pivot_table(df, values=values, index=indices[:n], columns=columns, aggfunc=aggfunc,
                                           fill_value=fill_value)
                table = table.reset_index()

                for column in indices[n:]:
                    table[column] = ''

                listOfTable.append(table)

            concatTable = pd.concat(listOfTable).sort_index()
            concatTable = concatTable.set_index(keys=indices)
            return concatTable.sort_index(axis=0, ascending=True)

        list_pivot_mese_entrate = ['pivot_gennaio_entrate',
                                   'pivot_febbraio_entrate',
                                   'pivot_marzo_entrate',
                                   'pivot_aprile_entrate',
                                   'pivot_maggio_entrate',
                                   'pivot_giugno_entrate',
                                   'pivot_luglio_entrate',
                                   'pivot_agosto_entrate',
                                   'pivot_settembre_entrate',
                                   'pivot_ottobre_entrate',
                                   'pivot_novembre_entrate',
                                   'pivot_dicembre_entrate'
                                   ]

        list_pivot_mese_uscite = ['pivot_gennaio_uscite',
                                  'pivot_febbraio_uscite',
                                  'pivot_marzo_uscite',
                                  'pivot_aprile_uscite',
                                  'pivot_maggio_uscite',
                                  'pivot_giugno_uscite',
                                  'pivot_luglio_uscite',
                                  'pivot_agosto_uscite',
                                  'pivot_settembre_uscite',
                                  'pivot_ottobre_uscite',
                                  'pivot_novembre_uscite',
                                  'pivot_dicembre_uscite'
                                  ]

        # pivot_table_w_subtotals
        #       (df=list_df_conti_camerino_mese_entrate[0],values='Euro',indices=['Entrate_Uscite', 'Categoria', 'Voce'],columns=[],aggfunc='sum',fill_value='')

        i = 0
        try:
            for x in range(12):
                list_pivot_mese_entrate[i] = np.round(pivot_table_w_subtotals
                                                  (df=list_df_conti_mese_entrate[i],
                                                   values='Euro',
                                                   indices=['Entrate_Uscite', 'Categoria', 'Voce'],
                                                   columns=[],
                                                   aggfunc='sum',
                                                   fill_value=''), 2)

                list_pivot_mese_uscite[i] = np.round(pivot_table_w_subtotals
                                                 (list_df_conti_mese_uscite[i],
                                                  values='Euro',
                                                  indices=['Entrate_Uscite', 'Categoria', 'Voce'],
                                                  columns=[],
                                                  aggfunc='sum',
                                                  fill_value=''), 2)
                i += 1
        except:
            pass



        list_pivot_mese_entrate = [list_pivot_mese_entrate[0],  # pivot gennaio entrate
                                   list_pivot_mese_entrate[1],
                                   list_pivot_mese_entrate[2],
                                   list_pivot_mese_entrate[3],
                                   list_pivot_mese_entrate[4],
                                   list_pivot_mese_entrate[5],
                                   list_pivot_mese_entrate[6],
                                   list_pivot_mese_entrate[7],
                                   list_pivot_mese_entrate[8],
                                   list_pivot_mese_entrate[9],
                                   list_pivot_mese_entrate[10],
                                   list_pivot_mese_entrate[11]
                                   ]

        list_pivot_mese_uscite = [list_pivot_mese_uscite[0],
                                  list_pivot_mese_uscite[1],
                                  list_pivot_mese_uscite[2],
                                  list_pivot_mese_uscite[3],
                                  list_pivot_mese_uscite[4],
                                  list_pivot_mese_uscite[5],
                                  list_pivot_mese_uscite[6],
                                  list_pivot_mese_uscite[7],
                                  list_pivot_mese_uscite[8],
                                  list_pivot_mese_uscite[9],
                                  list_pivot_mese_uscite[10],
                                  list_pivot_mese_uscite[11]
                                  ]

        # print(list_pivot_mese_entrate[0])

        ##############PRIMA PAGINA#############################

        # Creo il file 'conti_styled.xlsx'
        wb = Workbook()
        # La prima pagina 'Sheet' la chiamo 'Copertina_fronte'
        wb['Sheet'].title = ('Copertina_fronte')

        wb['Copertina_fronte'].merge_cells('A4:I4')
        wb['Copertina_fronte']['A4'] = 'Resoconto Amministrativo'
        wb['Copertina_fronte']['A4'].font = Font(name='Calibri',
                                                 size=35,
                                                 bold=True,
                                                 italic=True,
                                                 vertAlign='none',
                                                 underline='single',
                                                 strike=False,
                                                 color='204ac8')  # blu royal
        wb['Copertina_fronte']['A4'].alignment = Alignment(horizontal='center')

        wb['Copertina_fronte']['A7'] = 'Fraternità di .....'
        wb['Copertina_fronte']['A7'].font = Font(name='Calibri',
                                                 size=25,
                                                 bold=True,
                                                 italic=True,
                                                 vertAlign='none',
                                                 underline='single',
                                                 strike=False,
                                                 color='204ac8')

        wb['Copertina_fronte']['A10'] = 'Anno.....'
        wb['Copertina_fronte']['A10'].font = Font(name='Calibri',
                                                  size=25,
                                                  bold=True,
                                                  italic=True,
                                                  vertAlign='none',
                                                  underline='single',
                                                  strike=False,
                                                  color='204ac8')

        # Inserisco immagine bilancia
        img = openpyxl.drawing.image.Image('bilancia.png')
        img.anchor = 'B13'
        wb['Copertina_fronte'].add_image(img)

        wb.save('conti_styled.xlsx')
        ##################################

        # Con ExcelWriter di pandas METTO INSIEME il pivot delle entrate e il pivot delle uscite

        # sheets dei 12 mesi
        list_ws_mese = ['ws_gennaio',
                        'ws_febbraio',
                        'ws_marzo',
                        'ws_aprile',
                        'ws_maggio',
                        'ws_giugno',
                        'ws_luglio',
                        'ws_agosto',
                        'ws_settembre',
                        'ws_ottobre',
                        'ws_novembre',
                        'ws_dicembre'
                        ]

        c = 0  # contatore
        try:
            for x in range(12):
                with pd.ExcelWriter('conti_styled.xlsx',
                                    mode="a",
                                    engine="openpyxl",
                                    if_sheet_exists="overlay",
                                    ) as writer:
                    list_pivot_mese_entrate[c].to_excel(writer, sheet_name=list_mese[c], startrow=5)
                    list_pivot_mese_uscite[c].to_excel(writer, sheet_name=list_mese[c],
                                                    startrow=(len(list_pivot_mese_entrate[c]) + 10))

                # leggo il file "conti_camerino_styled.xlsx"
                wb = load_workbook(filename="conti_styled.xlsx")
                # creo i 12 sheet
                list_ws_mese[c] = wb[list_mese[c]]
                c += 1
        except:
            pass


        # sheets dei 12 mesi
        list_ws_mese = [wb[list_mese[0]],  # ws_gennaio,
                        wb[list_mese[1]],  # ws_febbraio,
                        wb[list_mese[2]],  # ws_marzo,
                        wb[list_mese[3]],  # ws_aprile,
                        wb[list_mese[4]],  # ws_maggio,
                        wb[list_mese[5]],  # ws_giugno,
                        wb[list_mese[6]],  # ws_luglio,
                        wb[list_mese[7]],  # ws_agosto,
                        wb[list_mese[8]],  # ws_settembre,
                        wb[list_mese[9]],  # ws_ottobre,
                        wb[list_mese[10]],  # ws_novembre,
                        wb[list_mese[11]],  # ws_dicembre
                        ]

        ################# APPLICO STILE ########################

        # Colonna D: Formattazione degli euro in valuta euro
        for sheet in list_ws_mese:
            for row in sheet[7:sheet.max_row]:  # skip the header
                # print(row) #(<Cell 'gennaio'.A7>, <Cell 'gennaio'.B7>, <Cell 'gennaio'.C7>, <Cell 'gennaio'.D7>)
                cell = row[3]  # il quarto valore della tuple
                print(cell)  # <Cell 'multiple'.D7>
                cell.number_format = '#,##0.00 €'
                cell.alignment = Alignment(horizontal="right")
                cell.font = Font(bold=True)

        # Aggiungo la scritta 'Totale =' alla tabella pivot davanti ai subtotali
        for sheet in list_ws_mese:
            for row in sheet[7:sheet.max_row]:
                for cell in sheet['B']:  # per ogni casella della colonna B
                    if cell.value is not None:
                        # ossia se la casella nella colonna B non è vuota
                        # Assegnale uno stile
                        cell.font = Font(name='Calibri',
                                         size=15,
                                         bold=True,
                                         italic=True,
                                         vertAlign='none',
                                         underline='single',
                                         strike=False,
                                         color='a81a1a')
                        # Assegna uno stile anche alla cell accanto corrispondente
                        sheet.cell(row=cell.offset(row=0, column=0).row, column=3,
                                   value=f"Totale {cell.value} =")
                        sheet.cell(row=cell.offset(row=0, column=0).row, column=4).font \
                            = Font(size=15, color='a81a1a', bold=True)
                        sheet.cell(row=cell.offset(row=0, column=0).row, column=4).number_format \
                            = '#,##0.00 €'
                        sheet.cell(row=cell.offset(row=0, column=0).row, column=4).alignment \
                            = Alignment(horizontal="left")

        # Larghezza fissa colonne
        i = 0  # contatore
        # set the height of the first row in each sheet
        for sheet in list_ws_mese:
            sheet.row_dimensions[1].height = 70

            # set the width of the column
            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 30
            sheet.column_dimensions['D'].width = 15

            # merge cells
            sheet.merge_cells('A1:D1')

            # scrivo nella cella 'A1'
            sheet['A1'].value = list_mese[i]
            i += 1

            # Formattazione cella
            sheet['A1'].font = Font(name='Calibri',
                                    size=25,
                                    bold=True,
                                    italic=True,
                                    vertAlign='none',
                                    underline='single',
                                    strike=False,
                                    color='a81a1a')

            sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

            # Colonna C: Allineamento
            for row in sheet[7:sheet.max_row]:  # skip the header
                cell = row[2]  # il terzo valore della tuple
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # Colonna D: Allineamento
            for row in sheet[7:sheet.max_row]:  # skip the header
                cell = row[1]  # il secondo valore della tuple
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Formattazione headers
            list = []

            for row in sheet.rows:
                for cell in row:
                    if (cell.value == ("Categoria") or
                            cell.value == ("Entrate") or
                            cell.value == ("Euro") or
                            cell.value == ("Uscite") or
                            cell.value == ("Voce")):
                        list.append(cell)
            for cell in list:
                cell.font = Font(name='Calibri', size=15, color='a81a1a', bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in sheet.rows:
                for cell in row:
                    if (cell.value == ("Totale Categoria =")):
                        cell.font = Font(name='Calibri', size=15, color='a81a1a', bold=True)
                        cell.alignment = Alignment(horizontal="right", vertical="center")

            # Formattazione ' Euro accanto a 'Totale categoria='

            double = Side(border_style="double", color="4617F1")
            cont = 0
            for row in sheet.rows:
                for cell in row:
                    if (cell.value == ('Totale Categoria =')):
                        #
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).font \
                            = Font(size=15, color='a81a1a', bold=True)
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).number_format \
                            = '#,##0.00 €'
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).alignment \
                            = Alignment(horizontal="left")
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).fill \
                            = PatternFill('solid', fgColor='d1d22e')
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).border \
                            = Border(bottom=double, top=double, left=double, right=double)

                        sheet.cell(row=cell.offset(row=1, column=1).row, column=4).font \
                            = Font(size=15, color='a81a1a', bold=True)
                        sheet.cell(row=cell.offset(row=1, column=1).row, column=4).number_format \
                            = '#,##0.00 €'
                        sheet.cell(row=cell.offset(row=1, column=1).row, column=4).alignment \
                            = Alignment(horizontal="left")
                        sheet.cell(row=cell.offset(row=1, column=1).row, column=4).fill \
                            = PatternFill('solid', fgColor='d1d22e')
                        sheet.cell(row=cell.offset(row=1, column=1).row, column=4).border \
                            = Border(bottom=double, top=double, left=double, right=double)

            # Rendi 'invisibile il testo"Entrate_Uscite"
            list = []

            # Formattazione "Entrate_Uscite", piccolo per non essere visto

            for row in sheet.rows:
                for cell in row:
                    if cell.value == ("Entrate_Uscite"):
                        list.append(cell)
            for cell in list:
                cell.font = Font(size=1)

        # imposto saldo iniziale
        # Passo da DoubleVar a float
        saldo = float(Report.saldo_report_func(saldo_report_DoubleVar))
        # imposto un contatore
        e = 0
        i = 0
        # memorizzo in una lista tutti i saldi dall'inizio alla fine dell'anno
        # le entrete
        # le uscite
        list_saldo_iniale_finale_anno = [saldo]
        list_entrate_mesi = []
        list_uscite_mesi = []
        # saldo iniziale
        for sheet in list_ws_mese:
            # coordinate dell'ultima cella della colonna A di ogni foglio

            last_cell_coordiate = 'C' + str(sheet.max_row)
            # print(last_cell_coordiate)
            # attraverso le coordinate risalgo alla cella di excel
            cell = sheet[last_cell_coordiate]
            # print(cell.value)

            # cell = sheet.cell(row=1, column=1)
            # last_cell = sheet[last_cell]
            #
            # print(cell.coordinate, cell.row, cell.column) # A184 184 1 per tutti e 12 i fogli
            #
            # #print(cell) # stampa ultima cella colonna A
            sheet.cell(row=cell.offset(row=5, column=0).row, column=2, value='SALDO del mese precedente').font \
                = Font(size=15, color='000000', bold=True)
            sheet.cell(row=cell.offset(row=5, column=0).row, column=2).alignment = Alignment(horizontal="left")
            sheet.cell(row=cell.offset(row=7, column=0).row, column=2, value='ENTRATE del mese').font \
                = Font(size=15, color='000000', bold=True)
            sheet.cell(row=cell.offset(row=7, column=0).row, column=2).alignment = Alignment(horizontal="left")
            sheet.cell(row=cell.offset(row=9, column=0).row, column=2, value='USCITE del mese').font \
                = Font(size=15, color='000000', bold=True)
            sheet.cell(row=cell.offset(row=9, column=0).row, column=2).alignment = Alignment(horizontal="left")
            sheet.cell(row=cell.offset(row=11, column=0).row, column=2, value='DIS/AVANZO del mese').font \
                = Font(size=15, color='000000', bold=True)
            sheet.cell(row=cell.offset(row=11, column=0).row, column=2).alignment = Alignment(horizontal="left")
            sheet.cell(row=cell.offset(row=13, column=0).row, column=2, value='SALDO del mese corrente').font \
                = Font(size=15, color='000000', bold=True)
            sheet.cell(row=cell.offset(row=13, column=0).row, column=2).alignment = Alignment(horizontal="left")

            sheet.cell(row=cell.offset(row=5, column=0).row, column=4,
                       value=   (
                           list_saldo_iniale_finale_anno[e]
                                )
                       )

            # mi calcolo il saldo finale e la assegno alla variabile saldo
            saldo = (list_saldo_iniale_finale_anno[e] +
                     list_df_conti_mese_entrate[e]['Euro'].sum(numeric_only=True) -
                     list_df_conti_mese_uscite[e]['Euro'].sum(numeric_only=True)
                     )

            # print(list_saldo_iniale_finale_anno)
            sheet.cell(row=cell.offset(row=5, column=0).row, column=4).font = Font(size=15, bold=True)
            sheet.cell(row=cell.offset(row=5, column=0).row, column=4).number_format = '#,##0.00€'
            sheet.cell(row=cell.offset(row=5, column=0).row, column=4).alignment = Alignment(horizontal="right")

            if sheet.cell(row=cell.offset(row=5, column=0).row, column=4).value > 0:
                sheet.cell(row=cell.offset(row=5, column=0).row, column=4).font = Font(color='000000', size=15,
                                                                                       bold=True)
            else:
                sheet.cell(row=cell.offset(row=5, column=0).row, column=4).font = Font(color='a81a1a', size=15,
                                                                                       bold=True)

            sheet.cell(row=cell.offset(row=7, column=0).row, column=4,
                       value=list_df_conti_mese_entrate[i]['Euro'].sum(numeric_only=True))

            sheet.cell(row=cell.offset(row=7, column=0).row, column=4).font = Font(size=15, color='000000',
                                                                                   bold=True)
            sheet.cell(row=cell.offset(row=7, column=0).row, column=4).number_format = '#,##0.00€'
            sheet.cell(row=cell.offset(row=7, column=0).row, column=4).alignment = Alignment(horizontal="right")
            sheet.cell(row=cell.offset(row=9, column=0).row, column=4,
                       value=list_df_conti_mese_uscite[i]['Euro'].sum(numeric_only=True))
            sheet.cell(row=cell.offset(row=9, column=0).row, column=4).font = Font(size=15, color='a81a1a',
                                                                                   bold=True)
            sheet.cell(row=cell.offset(row=9, column=0).row, column=4).number_format = '-#,##0.00€'
            sheet.cell(row=cell.offset(row=9, column=0).row, column=4).alignment = Alignment(horizontal="right")
            sheet.cell(row=cell.offset(row=11, column=0).row, column=4,
                       value=((list_df_conti_mese_entrate[i]['Euro']).sum(numeric_only=True) -
                              (list_df_conti_mese_uscite[i]['Euro']).sum(numeric_only=True)
                              )
                       )
            # sheet.cell(row=cell.offset(row=11, column=0).row, column=4).font = Font(size=15,
            #                                                                       bold=True)
            sheet.cell(row=cell.offset(row=11, column=0).row, column=4).number_format = '#,##0.00€'
            sheet.cell(row=cell.offset(row=11, column=0).row, column=4).alignment = Alignment(horizontal="right")

            if (sheet.cell(row=cell.offset(row=11, column=0).row, column=4).value) > 0:
                sheet.cell(row=cell.offset(row=11, column=0).row, column=4).font = Font(color='000000', size=15,
                                                                                        bold=True)
            else:
                sheet.cell(row=cell.offset(row=11, column=0).row, column=4).font = Font(color='a81a1a', size=15,

                                                                                        bold=True)
            # queste liste mi servono per il grafico
            list_saldo_iniale_finale_anno.append(saldo)
            list_entrate_mesi.append(list_df_conti_mese_entrate[i]['Euro'].sum(numeric_only=True))
            list_uscite_mesi.append(list_df_conti_mese_uscite[i]['Euro'].sum(numeric_only=True))

            # print(list_saldo_iniale_finale_anno)
            # print(list_entrate_mesi)
            # print(list_uscite_mesi)
            e += 1
            i += 1
            # Saldo finale
            # for row in sheet:
            #         for cell in row:
            #             if (cell.value == ("TOTALE_Uscite")):

            sheet.cell(row=cell.offset(row=13, column=0).row, column=4,
                       value=(
                           saldo
                       )
                       )

            sheet.cell(row=cell.offset(row=13, column=0).row, column=4).font = Font(size=15, bold=True)
            sheet.cell(row=cell.offset(row=13, column=0).row, column=4).number_format = '#,##0.00€'
            sheet.cell(row=cell.offset(row=13, column=0).row, column=4).alignment = Alignment(horizontal="right")

            if sheet.cell(row=cell.offset(row=13, column=0).row, column=4).value > 0:
                sheet.cell(row=cell.offset(row=13, column=0).row, column=4).font = Font(color='000000', size=15,
                                                                                        bold=True)
            else:
                sheet.cell(row=cell.offset(row=13, column=0).row, column=4).font = Font(color='a81a1a', size=15,
                                                                                        bold=True)

        ###################### tabellone entrate
        import openpyxl
        from openpyxl.worksheet import page

        # Creo un nuovo foglio
        ws_tab_entrate = wb.create_sheet('Tab_Entrate')
        ws_tab_entrate.set_printer_settings(Worksheet.PAPERSIZE_A4, Worksheet.ORIENTATION_LANDSCAPE)
        # ws_tab_entrate.print_area = 'A1:Z1'

        ws_tab_entrate['A1'] = 'Tabellone Entrate'
        ws_tab_entrate.row_dimensions[1].height = 70
        ws_tab_entrate['A1'].font = Font(name='Calibri', size=80, color='a81a1a', bold=True)
        ws_tab_entrate['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws_tab_entrate.merge_cells('A1:N1')

        # set the width of the column
        ws_tab_entrate.column_dimensions['A'].width = 14
        ws_tab_entrate.column_dimensions['B'].width = 8
        ws_tab_entrate.column_dimensions['C'].width = 8
        ws_tab_entrate.column_dimensions['D'].width = 8
        ws_tab_entrate.column_dimensions['E'].width = 8
        ws_tab_entrate.column_dimensions['F'].width = 8
        ws_tab_entrate.column_dimensions['G'].width = 8
        ws_tab_entrate.column_dimensions['H'].width = 8
        ws_tab_entrate.column_dimensions['I'].width = 8
        ws_tab_entrate.column_dimensions['L'].width = 9
        ws_tab_entrate.column_dimensions['M'].width = 8
        ws_tab_entrate.column_dimensions['N'].width = 14

        i = 2
        for x in range(2, 20):
            ws_tab_entrate.row_dimensions[i].height = 20
            i += 1

        # anno =
        df_conti_camerino_TOT_entrate = df_database_conti.loc[
            (df_database_conti['Anno'] == Report.anno_report_func(anno_report_IntVar)) &
            (df_database_conti['Entrate_Uscite'] == 'Entrate')]
        # print(df_conti_camerino_TOT_entrate.head(40))

        pivot_conti_camerino_TOT_entrate = np.round(pd.pivot_table
                                                    (df_conti_camerino_TOT_entrate,
                                                     values='Euro',
                                                     index=['Categoria'],
                                                     columns='Mese',
                                                     aggfunc='sum',
                                                     margins=True,
                                                     margins_name='TOTALE_Entrate',
                                                     fill_value=0), 2)

        # print(pivot_conti_camerino_TOT_entrate.head())

        for r in dataframe_to_rows(pivot_conti_camerino_TOT_entrate, index=True, header=True):
            ws_tab_entrate.append(r)

        for cell in ws_tab_entrate['A'] + ws_tab_entrate[2]:
            cell.style = 'Pandas'

        ws_tab_entrate['A1'].font = Font(name='Calibri', size=40, color='a81a1a', bold=True)
        ws_tab_entrate['A1'].alignment = Alignment(horizontal="center", vertical="center")

        TOTALE_ENTRATE = round(df_conti_camerino_TOT_entrate['Euro'].sum(), 2)

        for row in ws_tab_entrate.rows:
            for cell in row:
                if cell.value == TOTALE_ENTRATE:
                    cell.font = Font(name='Calibri', size=13, color='000000', bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = Border(bottom=double, top=double, left=double, right=double)
                    cell.fill = PatternFill('solid', fgColor='d1d22e')
                    cell.number_format = '#,##0.00€'

            # list_df_conti_camerino_mese_uscite[i] = df_conti_camerino_modified.loc[
            #     (df_conti_camerino_modified['Anno'] == anno) &
            #     (df_conti_camerino_modified['Mese'] == list_mese[i]) &
            #     (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]
            # #print(list_df_conti_camerino_mese_uscite[i].head())
            #
            # i += 1

        ###################### tabellone uscite
        # Creo un nuovo foglio
        ws_tab_uscite = wb.create_sheet('Tab_Uscite')

        ws_tab_uscite.set_printer_settings(Worksheet.PAPERSIZE_A4, Worksheet.ORIENTATION_LANDSCAPE)

        ws_tab_uscite['A1'] = 'Tabellone Uscite'
        ws_tab_uscite.merge_cells('A1:N1')
        ws_tab_uscite.row_dimensions[1].height = 45

        df_conti_camerino_TOT_uscite = df_database_conti.loc[
            (df_database_conti['Anno'] == Report.anno_report_func(anno_report_IntVar)) &
            (df_database_conti['Entrate_Uscite'] == 'Uscite')]

        # print(df_conti_camerino_TOT_uscite.head(40))
        pivot_conti_camerino_TOT_uscite = np.round(pd.pivot_table
                                                   (df_conti_camerino_TOT_uscite,
                                                    values='Euro',
                                                    # index=['Entrate_Uscite', 'Categoria', 'Voce'],
                                                    index=['Categoria'],
                                                    columns='Mese',
                                                    aggfunc='sum',
                                                    margins=True,
                                                    margins_name='TOTALE_Uscite',
                                                    fill_value=0), 2)

        print(pivot_conti_camerino_TOT_uscite.head())
        for r in dataframe_to_rows(pivot_conti_camerino_TOT_uscite, index=True, header=True):
            ws_tab_uscite.append(r)

        for cell in ws_tab_uscite['A'] + ws_tab_uscite[2]:
            cell.style = 'Pandas'

        # set the width of the column
        ws_tab_uscite.column_dimensions['A'].width = 16
        ws_tab_uscite.column_dimensions['B'].width = 7
        ws_tab_uscite.column_dimensions['C'].width = 7
        ws_tab_uscite.column_dimensions['D'].width = 7
        ws_tab_uscite.column_dimensions['E'].width = 7
        ws_tab_uscite.column_dimensions['F'].width = 7
        ws_tab_uscite.column_dimensions['G'].width = 7
        ws_tab_uscite.column_dimensions['H'].width = 7
        ws_tab_uscite.column_dimensions['I'].width = 7
        ws_tab_uscite.column_dimensions['L'].width = 9
        ws_tab_uscite.column_dimensions['M'].width = 7
        ws_tab_uscite.column_dimensions['N'].width = 16

        i = 2
        for x in range(2, 20):
            ws_tab_uscite.row_dimensions[i].height = 17
            i += 1

        ws_tab_uscite['A1'].font = Font(name='Calibri', size=40, color='a81a1a', bold=True)
        ws_tab_uscite['A1'].alignment = Alignment(horizontal="center", vertical="center")

        TOTALE_USCITE = round(df_conti_camerino_TOT_uscite['Euro'].sum(), 2)

        for row in ws_tab_uscite.rows:
            for cell in row:
                if cell.value == TOTALE_USCITE:
                    cell.font = Font(name='Calibri', size=13, color='000000', bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = Border(bottom=double, top=double, left=double, right=double)
                    cell.fill = PatternFill('solid', fgColor='d1d22e')
                    cell.number_format = '#,##0.00€'

        ######################  grafico

        from openpyxl.chart import Reference, LineChart

        # Creo un nuovo foglio
        ws_saldo_riepilogo = wb.create_sheet('Saldo_riepilogo')
        ws_saldo_riepilogo['A1'] = 'Tabella Entrate - Uscite - Saldo di ogni mese '

        # Formattazione
        ws_saldo_riepilogo.row_dimensions[1].height = 100
        ws_saldo_riepilogo.merge_cells('A1:E1')

        # set the width of the column
        ws_saldo_riepilogo.column_dimensions['A'].width = 15
        ws_saldo_riepilogo.column_dimensions['B'].width = 17
        ws_saldo_riepilogo.column_dimensions['C'].width = 17
        ws_saldo_riepilogo.column_dimensions['D'].width = 17
        ws_saldo_riepilogo.column_dimensions['E'].width = 17

        # queste liste mi servono per il grafico

        list_headers = ['mese', 'saldo_iniziale', 'entrate_mese', 'uscite_mese', 'saldo_finale']
        list_saldo_iniziale_anno = list_saldo_iniale_finale_anno[:-1]
        # list_entrate_mesi
        # list_uscite_mesi
        list_saldo_finale_anno = list_saldo_iniale_finale_anno[1:]

        i = 0
        ws_saldo_riepilogo.append(list_headers)
        for mese in range(1, 13):
            mese_saldo_grafico = [list_mese[i], list_saldo_iniziale_anno[i], list_entrate_mesi[i], list_uscite_mesi[i],
                                  list_saldo_finale_anno[i]]

            ws_saldo_riepilogo.append(mese_saldo_grafico)
            i += 1

        list = []
        for row in ws_saldo_riepilogo.rows:
            for cell in row:

                if (cell.value == ("saldo_iniziale") or
                        cell.value == ("entrate_mese") or
                        cell.value == ("uscite_mese") or
                        cell.value == ("saldo_finale") or
                        cell.value == ("gennaio") or
                        cell.value == ("febbraio") or
                        cell.value == ("marzo") or
                        cell.value == ("aprile") or
                        cell.value == ("maggio") or
                        cell.value == ("giugno") or
                        cell.value == ("luglio") or
                        cell.value == ("agosto") or
                        cell.value == ("settembre") or
                        cell.value == ("ottobre") or
                        cell.value == ("novembre") or
                        cell.value == ("dicembre")
                ):
                    cell.font = Font(name='Calibri', size=13, color='000000', bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            for cell in row:
                if cell.value == ("mese"):
                    cell.font = Font(size=1)

        for row in ws_saldo_riepilogo.iter_rows(min_row=3, min_col=2, max_row=14, max_col=5):
            for cell in row:
                cell.font = Font(name='Calibri', size=13, color='000000', bold=True)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00 €'
                if (int(cell.value) > 0):
                    cell.font = Font(color='000000')
                else:
                    cell.font = Font(color='a81a1a')

        ws_saldo_riepilogo['A1'].font = Font(name='Calibri', size=20, color='a81a1a', bold=True)
        ws_saldo_riepilogo['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws_saldo_riepilogo['B3'].border \
            = Border(bottom=double, top=double, left=double, right=double)
        ws_saldo_riepilogo['B3'].fill \
            = PatternFill('solid', fgColor='d1d22e')
        ws_saldo_riepilogo['E14'].border \
            = Border(bottom=double, top=double, left=double, right=double)
        ws_saldo_riepilogo['E14'].fill \
            = PatternFill('solid', fgColor='d1d22e')

        data = Reference(ws_saldo_riepilogo, min_col=3, min_row=2, max_col=5, max_row=14)
        titles = Reference(ws_saldo_riepilogo, min_row=3, max_row=14, min_col=1)

        chart = LineChart()
        chart.title = "Bilancio"
        chart.style = 12

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = 'Mesi'
        chart.y_axis.title = 'Euro'

        ws_saldo_riepilogo.add_chart(chart, "A21")
        #######################PAGINA CONCLUSIVA
        # format
        from openpyxl.styles import NamedStyle, Font, Border, Side
        # HIGHLIGHT
        highlight = NamedStyle(name="highlight")
        highlight.font = Font(name='Calibri', size=15, color='000000', bold=True)
        double = Side(border_style="double", color="4617F1")
        # highlight.border = Border(bottom=double, top=double, left=double, right=double)
        highlight.fill = PatternFill('solid', fgColor='d1d22e')
        highlight.alignment = Alignment(horizontal="right", vertical="center")
        highlight.number_format = '#,##0.00 €'
        wb.add_named_style(highlight)

        ############BLACK
        black = NamedStyle(name="black")
        black.font = Font(name='Calibri', size=15, color='000000', bold=True)
        black.alignment = Alignment(horizontal="right", vertical="center")
        wb.add_named_style(black)

        import openpyxl
        from openpyxl import load_workbook
        ws_fine = wb.create_sheet('Fine')
        ws_fine['A1'] = 'Bilancio Anno ...'
        ws_fine.merge_cells('A1:I1')
        ws_fine.row_dimensions[1].height = 45
        ws_fine['A1'].font = Font(name='Calibri', size=35, color='a81a1a', bold=True)
        ws_fine['A1'].alignment = Alignment(horizontal="center", vertical="center")

        ws_fine['A7'] = 'RIEPILOGO'
        ws_fine.merge_cells('A7:C7')

        ws_fine['A10'] = 'SALDO iniziale'
        ws_fine.merge_cells('A10:C10')
        ws_fine['E10'] = saldo_report_DoubleVar.get()
        ws_fine.merge_cells('E10:G10')
        ws_fine['E10'].style = 'highlight'

        ws_fine['A13'] = 'TOTALE Entrate'
        ws_fine.merge_cells('A13:C13')
        ws_fine['E13'] = TOTALE_ENTRATE
        ws_fine.merge_cells('E13:G13')
        ws_fine['E13'].style = 'highlight'

        ws_fine['A16'] = 'TOTALE Uscite'
        ws_fine.merge_cells('A16:C16')
        ws_fine['E16'] = TOTALE_USCITE
        ws_fine.merge_cells('E16:G16')
        ws_fine['E16'].style = 'highlight'

        ws_fine['A19'] = 'DIS/AVANZO'
        ws_fine.merge_cells('A19:C19')
        ws_fine['E19'] = (TOTALE_ENTRATE - TOTALE_USCITE)
        ws_fine.merge_cells('E19:G19')
        ws_fine['E19'].style = 'highlight'

        ws_fine['A22'] = 'SALDO Finale'
        ws_fine.merge_cells('A22:C22')
        ws_fine['E22'] = saldo
        ws_fine.merge_cells('E22:G22')
        ws_fine['E22'].style = 'highlight'

        ws_fine['A7'].font = Font(name='Calibri', size=20, color='a81a1a', bold=True)
        ws_fine['A7'].alignment = Alignment(horizontal="center", vertical="center")
        ws_fine['A10'].font = Font(name='Calibri', size=15, bold=True)
        ws_fine['A10'].alignment = Alignment(horizontal="right", vertical="center")
        ws_fine['A13'].font = Font(name='Calibri', size=15, bold=True)
        ws_fine['A13'].alignment = Alignment(horizontal="right", vertical="center")
        ws_fine['A16'].font = Font(name='Calibri', size=15, bold=True)
        ws_fine['A16'].alignment = Alignment(horizontal="right", vertical="center")
        ws_fine['A19'].font = Font(name='Calibri', size=15, bold=True)
        ws_fine['A19'].alignment = Alignment(horizontal="right", vertical="center")
        ws_fine['A22'].font = Font(name='Calibri', size=15, bold=True)
        ws_fine['A22'].alignment = Alignment(horizontal="right", vertical="center")

        ws_fine['A32'] = 'Data'
        ws_fine['A32'].style = 'black'
        ws_fine.merge_cells('A32:B32')
        ws_fine['D29'] = 'Guardiano'
        ws_fine['D29'].style = 'black'
        ws_fine.merge_cells('D29:I29')
        ws_fine['d35'] = 'Vicario'
        ws_fine['D35'].style = 'black'
        ws_fine.merge_cells('D35:I35')
        ws_fine['D41'] = 'Economo'
        ws_fine['D41'].style = 'black'
        ws_fine.merge_cells('D41:I41')
        ws_fine['A38'] = 'Timbro'
        ws_fine['A38'].style = 'black'
        ws_fine.merge_cells('A38:B38')



        wb.save('conti_styled.xlsx')

        if sys.platform == "win32":
            os.startfile('conti_styled.xlsx')
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, 'conti_styled.xlsx'])
      except:
            messagebox.showwarning(title='Dati Mancanti', message="L'anno indicato deve avere almeno una voce in entrata e una voce in uscita")





'''

# B_add = Button(Frame1in, text='add', width=10, command=lambda:[submit()]).grid(row=0, column=0, padx=20, pady=15)

B_add = Button(Frame_button_in_combo, text='aggiungi', width=10, command=lambda: [submit(), query_database()]).pack(side=RIGHT, padx=10)

B_update = Button(Frame_button_in_update, text='aggiorna', width=10, command=update_record).pack(side=RIGHT, padx=10)
B_delete = Button(Frame_button_in_update, text='cancella', width=10, command=remove_one).pack(side=RIGHT, padx=10)


B_excel = Button(Frame_button_in_Treeview, text='Filtra con excel', width=10, command=sqlite3_to_excel).pack(side=RIGHT, padx=10)

'''
# B_report = Button(Frame_excell_botton, text='report', width=10, command= lambda: print(Report.anno_report_func(anno_report_Stringvar))).grid(row=0, column=2, padx=20, pady=15)
B_report = Button(Frame_excell_botton, text='report', width=10,
                  command=lambda: Report.report(Report.anno_report_func(anno_report_IntVar))).grid(row=0, column=2,
                                                                                                      padx=20, pady=15)

# B_report = Button(Frame_excell_botton, text='report', width=10, command=lambda: Report.report(2012)).grid(row=0, column=2, padx=20, pady=15)
B_esporta = Button(Frame_excell_botton, text='esporta', width=10, command='').grid(row=0, column=3, padx=20, pady=15)
B_importa = Button(Frame_excell_botton, text='importa', width=10, command='').grid(row=0, column=4, padx=20, pady=15)
#####


query_database()

# conn.close()
'''

query_database()

root.mainloop()
