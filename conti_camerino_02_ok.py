import pandas as pd

import numpy as np

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Leggi il file xlsx e trasformalo in dataframe impostando i nomi colonna
col_names = ['Anno', 'Mese', 'Categoria', 'Voce', 'Euro']
df_conti_camerino_modified = pd.read_excel('conti_camerino_da_importare.xlsx', names=col_names)

# Creo la colonna ['Entrate_Uscite'] e scrivo in automatico i volori
df_conti_camerino_modified['Entrate_Uscite'] = df_conti_camerino_modified['Categoria'].apply \
    (lambda x: 'Entrate' if x == 'Vendite varie' or
                            x == 'Salute' or
                            x == 'Curia' or
                            x == 'Collette-Chiesa' or
                            x == 'Congrua' or
                            x == 'Interessi' or
                            x == 'Messe celebrate' or
                            x == 'Offerte' or
                            x == 'Pensioni' or
                            x == 'Predicazione' or
                            x == 'Servizi religiosi' or
                            x == 'Stipendi' or
                            x == 'Sussidi' or
                            x == 'Rimbosi' or
                            x == 'Vendite varie' or
                            x == 'Eccedenza Cassa'
    else 'Uscite')

# Assegno alle colonne l'ordine desiderato
df_conti_camerino_modified = df_conti_camerino_modified[['Anno', 'Mese', 'Entrate_Uscite', 'Categoria', 'Voce', 'Euro']]
df_conti_camerino_modified["Mese"] = df_conti_camerino_modified["Mese"].astype("category")
df_conti_camerino_modified["Mese"] = df_conti_camerino_modified["Mese"].cat.set_categories(
    ["gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno", "luglio", "agosto", "settembre", "ottobre",
     "novembre", "dicembre"])



# sheets dei 12 mesi
list_df_conti_camerino_mese_entrate = [ 'df_conti_camerino_mese_entrate_gennaio',
                                         'df_conti_camerino_mese_entrate_febbraio',
                                         'df_conti_camerino_mese_entrate_marzo',
                                         'df_conti_camerino_mese_entrate_aprile',
                                         'df_conti_camerino_mese_entrate_maggio',
                                         'df_conti_camerino_mese_entrate_giugno',
                                         'df_conti_camerino_mese_entrate_luglio',
                                         'df_conti_camerino_mese_entrate_agosto',
                                         'df_conti_camerino_mese_entrate_settembre',
                                         'df_conti_camerino_mese_entrate_ottobre',
                                         'df_conti_camerino_mese_entrate_novembre',
                                         'df_conti_camerino_mese_entrate_dicembre'
                                         ]

list_df_conti_camerino_mese_uscite = [ 'df_conti_camerino_mese_uscite_gennaio',
                                        'df_conti_camerino_mese_uscite_febbraio',
                                        'df_conti_camerino_mese_uscite_marzo',
                                        'df_conti_camerino_mese_uscite_aprile',
                                        'df_conti_camerino_mese_uscite_maggio',
                                        'df_conti_camerino_mese_uscite_giugno',
                                        'df_conti_camerino_mese_uscite_luglio',
                                        'df_conti_camerino_mese_uscite_agosto',
                                        'df_conti_camerino_mese_uscite_settembre',
                                        'df_conti_camerino_mese_uscite_ottobre',
                                        'df_conti_camerino_mese_uscite_novembre',
                                        'df_conti_camerino_mese_uscite_dicembre'
                                        ]

list_mese = ['gennaio',
             'febbraio',
             'marzo',
             'aprile',
             'maggio',
             'giugno',
             'luglio',
             'agosto',
             'settembre',
             'ottobre',
             'novembre',
             'dicembre'
             ]


#imposto anno
anno = 2015
i = 0
for x in range (12):
        list_df_conti_camerino_mese_entrate[i]= df_conti_camerino_modified.loc[
            (df_conti_camerino_modified['Anno'] == anno) &
            (df_conti_camerino_modified['Mese'] == list_mese[i]) &
            (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]
        #print(list_df_conti_camerino_pivot_entrate[i].head())


        list_df_conti_camerino_mese_uscite[i] = df_conti_camerino_modified.loc[
            (df_conti_camerino_modified['Anno'] == anno) &
            (df_conti_camerino_modified['Mese'] == list_mese[i]) &
            (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]
        #print(list_df_conti_camerino_mese_uscite[i].head())

        i += 1


list_df_conti_camerino_mese_entrate = [  list_df_conti_camerino_mese_entrate[0], #df_conti_camerino_pivot_entrate_gennaio
                                            list_df_conti_camerino_mese_entrate[1],
                                            list_df_conti_camerino_mese_entrate[2],
                                            list_df_conti_camerino_mese_entrate[3],
                                            list_df_conti_camerino_mese_entrate[4],
                                            list_df_conti_camerino_mese_entrate[5],
                                            list_df_conti_camerino_mese_entrate[6],
                                            list_df_conti_camerino_mese_entrate[7],
                                            list_df_conti_camerino_mese_entrate[8],
                                            list_df_conti_camerino_mese_entrate[9],
                                            list_df_conti_camerino_mese_entrate[10],
                                            list_df_conti_camerino_mese_entrate[11]
                                        ]

list_df_conti_camerino_mese_uscite = [   list_df_conti_camerino_mese_uscite[0], #df_conti_camerino_pivot_uscite_gennaio
                                            list_df_conti_camerino_mese_uscite[1],
                                            list_df_conti_camerino_mese_uscite[2],
                                            list_df_conti_camerino_mese_uscite[3],
                                            list_df_conti_camerino_mese_uscite[4],
                                            list_df_conti_camerino_mese_uscite[5],
                                            list_df_conti_camerino_mese_uscite[6],
                                            list_df_conti_camerino_mese_uscite[7],
                                            list_df_conti_camerino_mese_uscite[8],
                                            list_df_conti_camerino_mese_uscite[9],
                                            list_df_conti_camerino_mese_uscite[10],
                                            list_df_conti_camerino_mese_uscite[11]
                                        ]


list_pivot_mese_entrate = ['pivot_gennaio_entrate',
                              'pivot_febbraio_entrate',
                              'pivot_marzo_entrate',
                              'pivot_aprile_entrate',
                              'pivot_maggio_entrate',
                              'pivot_giugno_entrate',
                              'pivot_luglio_entrate',
                              'pivot_agosto_entrate',
                              'pivot_settembre_entrate',
                              'pivot_ottobre_entrate',
                              'pivot_novembre_entrate',
                              'pivot_dicembre_entrate'
                              ]

list_pivot_mese_uscite = ['pivot_gennaio_uscite',
                              'pivot_febbraio_uscite',
                              'pivot_marzo_uscite',
                              'pivot_aprile_uscite',
                              'pivot_maggio_uscite',
                              'pivot_giugno_uscite',
                              'pivot_luglio_uscite',
                              'pivot_agosto_uscite',
                              'pivot_settembre_uscite',
                              'pivot_ottobre_uscite',
                              'pivot_novembre_uscite',
                              'pivot_dicembre_uscite'
                              ]

i=0
for x in range(12):
    list_pivot_mese_entrate[i] = np.round(pd.pivot_table
                                 (list_df_conti_camerino_mese_entrate[i],
                                  values='Euro',
                                  index=['Entrate_Uscite', 'Categoria', 'Voce'],
                                  aggfunc='sum',
                                  margins=True,
                                  margins_name='TOTALE_Entrate',
                                  fill_value=0), 2)

    list_pivot_mese_uscite[i] = np.round(pd.pivot_table
                                 (list_df_conti_camerino_mese_uscite[i],
                                  values='Euro',
                                  index=['Entrate_Uscite', 'Categoria', 'Voce'],
                                  aggfunc='sum',
                                  margins=True,
                                  margins_name='TOTALE_Uscite',
                                  fill_value=0), 2)
    i +=1

list_pivot_mese_entrate=[list_pivot_mese_entrate[0], #pivot gennaio entrate
                    list_pivot_mese_entrate[1],
                    list_pivot_mese_entrate[2],
                    list_pivot_mese_entrate[3],
                    list_pivot_mese_entrate[4],
                    list_pivot_mese_entrate[5],
                    list_pivot_mese_entrate[6],
                    list_pivot_mese_entrate[7],
                    list_pivot_mese_entrate[8],
                    list_pivot_mese_entrate[9],
                    list_pivot_mese_entrate[10],
                    list_pivot_mese_entrate[11]
                    ]

list_pivot_mese_uscite= [list_pivot_mese_uscite[0],
                    list_pivot_mese_uscite[1],
                    list_pivot_mese_uscite[2],
                    list_pivot_mese_uscite[3],
                    list_pivot_mese_uscite[4],
                    list_pivot_mese_uscite[5],
                    list_pivot_mese_uscite[6],
                    list_pivot_mese_uscite[7],
                    list_pivot_mese_uscite[8],
                    list_pivot_mese_uscite[9],
                    list_pivot_mese_uscite[10],
                    list_pivot_mese_uscite[11]
                    ]


# Creo il file 'conti_camerino_styled.xlsx'
wb = Workbook()
wb['Sheet'].title = 'Copertina_fronte'
wb.save('conti_camerino_styled.xlsx')

# Con ExcelWriter di pandas metto insieme il pivot delle entrate ed il pivot delle uscite
list_mese = ['gennaio',
             'febbraio',
             'marzo',
             'aprile',
             'maggio',
             'giugno',
             'luglio',
             'agosto',
             'settembre',
             'ottobre',
             'novembre',
             'dicembre'
             ]


# sheets dei 12 mesi
list_ws_mese = ['ws_gennaio',
                'ws_febbraio',
                'ws_marzo',
                'ws_aprile',
                'ws_maggio',
                'ws_giugno',
                'ws_luglio',
                'ws_agosto',
                'ws_settembre',
                'ws_ottobre',
                'ws_novembre',
                'ws_dicembre'
                ]

c = 0 # contatore

for x in range (12):
    with pd.ExcelWriter('conti_camerino_styled.xlsx',
                        mode="a",
                        engine="openpyxl",
                        if_sheet_exists="overlay",
                        ) as writer:
        list_pivot_mese_entrate[c].to_excel(writer, sheet_name=list_mese[c], startrow=5)
        list_pivot_mese_uscite[c].to_excel(writer, sheet_name=list_mese[c], startrow=(len(list_pivot_mese_entrate[c] ) + 10))

    # leggo il file "conti_camerino_styled.xlsx"
    wb = load_workbook(filename="conti_camerino_styled.xlsx")
    #creo i 12 sheet
    list_ws_mese[c]= wb[list_mese[c]]

    c += 1



#sheets dei 12 mesi
list_ws_mese = [wb[list_mese[0]],  #ws_gennaio,
                wb[list_mese[1]],  #ws_febbraio,
                wb[list_mese[2]],  #ws_marzo,
                wb[list_mese[3]],  #ws_aprile,
                wb[list_mese[4]],  #ws_maggio,
                wb[list_mese[5]],  #ws_giugno,
                wb[list_mese[6]],  #ws_luglio,
                wb[list_mese[7]],  #ws_agosto,
                wb[list_mese[8]],  #ws_settembre,
                wb[list_mese[9]],  #ws_ottobre,
                wb[list_mese[10]],  #ws_novembre,
                wb[list_mese[11]],  #ws_dicembre
                ]


# list_mese = ['Conto del mese di Gennaio',
#              'Conto del mese di Febbraio',
#              'Conto del mese di Marzo',
#              'Conto del mese di Aprile',
#              'Conto del mese di Maggio',
#              'Conto del mese di Giugno',
#              'Conto del mese di Luglio',
#              'Conto del mese di Agosto',
#              'Conto del mese di Settembre',
#              'Conto del mese di Ottobre',
#              'Conto del mese di Novembre',
#              'Conto del mese di Dicembre',
#              ]

i = 0  # contatore
# set the height of the first row in each sheet
for sheet in list_ws_mese:
    sheet.row_dimensions[1].height = 70

    # set the width of the column
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 15

    # merge cells
    sheet.merge_cells('A1:D1')

    # scrivo nella cella 'A1'
    sheet['A1'].value = list_mese[i]
    i += 1

    # Formattazione cella
    sheet['A1'].font = Font(name='Calibri',
                            size=25,
                            bold=True,
                            italic=True,
                            vertAlign='none',
                            underline='single',
                            strike=False,
                            color='a81a1a')

    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

# Colonna D :Formattazione degli euro in valuta euro
#for sheet in list_ws_mese:
    for row in sheet[7:sheet.max_row]:  # skip the header
        # print(row) #(<Cell 'multiple'.A7>, <Cell 'multiple'.B7>, <Cell 'multiple'.C7>, <Cell 'multiple'.D7>)
        cell = row[3]  # il quarto valore della tuple
        # print (cell)# <Cell 'multiple'.D7>
        cell.number_format = '#,##0.00€'
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(bold=True)

    # Colonna C: Allineamento
    for row in sheet[7:sheet.max_row]:  # skip the header
        cell = row[2]  # il terzo valore della tuple
        cell.alignment = Alignment(horizontal="right")

    # Colonna D: Allineamento
    for row in sheet[7:sheet.max_row]:  # skip the header
        cell = row[1]  # il secondo valore della tuple
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Formattazione headers
    list = []

    for row in sheet.rows:
        for cell in row:
            if (cell.value == ("Categoria") or
                cell.value == ("Entrate") or
                cell.value == ("Euro") or
                cell.value == ("Uscite") or
                cell.value == ("Voce")):
                list.append(cell)
    for cell in list:
        cell.font = Font(size=15, color='a81a1a', bold=True)

    # Formattazione 'TOTALE_Entrate' e 'TOTALE_Uscite'
    list = []

    for row in sheet.rows:
        for cell in row:
            if (cell.value == ('TOTALE_Entrate') or
                    cell.value == ('TOTALE_Uscite')):
                list.append(cell)
        for cell in list:
            cell.font = Font(size=12, color='a81a1a', bold=True)

    # Rendi 'invisibile il testo"Entrate_Uscite"
    list = []

    # Formattazione "Entrate_Uscite"

    for row in sheet.rows:
        for cell in row:
            if cell.value == ("Entrate_Uscite"):
                list.append(cell)
    for cell in list:
        cell.font = Font(size=1)

    # for cell in list:
    #     cell.font = Font(size=1)
    #     print(cell)
    #     print(cell.coordinate)
    #     print(cell.row)
    #     print(cell.column)

    # Formattazione Euro somma totale
    list = []

    for row in sheet.rows:
        for cell in row:
            if (cell.value == ("TOTALE_Entrate") or
                cell.value == ("TOTALE_Uscite")):
                list.append(cell)
    for cell in list:
        sheet.cell(cell.row, column=4).font = Font(size=15, color='a81a1a', bold=True)

    # Creo tabella SAlDO


    for row in sheet.rows:
        for cell in row:
            if (cell.value == ("TOTALE_Uscite")):
                # print(cell.coordinate, cell.row, cell.column) # A184 184 1 per tutti e 12 i fogli
                # print(sheet.cell(row=cell.row, column=cell.column).value) # stampa 'TOTALE_Uscite
                sheet.cell(row=cell.offset(row=5, column=0).row, column=2, value='SALDO del mese precedente').font \
                    = Font(size=15, color='000000', bold=True)
                sheet.cell(row=cell.offset(row=5, column=0).row, column=2).alignment = Alignment(horizontal="left")
                sheet.cell(row=cell.offset(row=7, column=0).row, column=2, value='ENTRATE del mese').font \
                    = Font(size=15, color='000000', bold=True)
                sheet.cell(row=cell.offset(row=7, column=0).row, column=2).alignment = Alignment(horizontal="left")
                sheet.cell(row=cell.offset(row=9, column=0).row, column=2, value='USCITE del mese').font \
                    = Font(size=15, color='000000', bold=True)
                sheet.cell(row=cell.offset(row=9, column=0).row, column=2).alignment = Alignment(horizontal="left")
                sheet.cell(row=cell.offset(row=11, column=0).row, column=2, value='DIS/AVANZO del mese').font \
                    = Font(size=15, color='000000', bold=True)
                sheet.cell(row=cell.offset(row=11, column=0).row, column=2).alignment = Alignment(horizontal="left")
                sheet.cell(row=cell.offset(row=13, column=0).row, column=2, value='SALDO del mese corrente').font \
                    = Font(size=15, color='000000', bold=True)
                sheet.cell(row=cell.offset(row=13, column=0).row, column=2).alignment = Alignment(horizontal="left")



    # sheets dei 12 mesi


# list_df_conti_entrate = [list_df_conti_camerino_pivot_entrate[0],  # df_conti_camerino_pivot_entrate_gennaio
#                           list_df_conti_camerino_pivot_entrate[1],
#                           list_df_conti_camerino_pivot_entrate[2],
#                           list_df_conti_camerino_pivot_entrate[3],
#                           list_df_conti_camerino_pivot_entrate[4],
#                           list_df_conti_camerino_pivot_entrate[5],
#                           list_df_conti_camerino_pivot_entrate[6],
#                           list_df_conti_camerino_pivot_entrate[7],
#                           list_df_conti_camerino_pivot_entrate[8],
#                           list_df_conti_camerino_pivot_entrate[9],
#                           list_df_conti_camerino_pivot_entrate[10],
#                           list_df_conti_camerino_pivot_entrate[11]
#                           ]
#
# list_df_conti_uscite = [list_df_conti_camerino_pivot_uscite[0],  # df_conti_camerino_pivot_uscite_gennaio
#                          list_df_conti_camerino_pivot_uscite[1],
#                          list_df_conti_camerino_pivot_uscite[2],
#                          list_df_conti_camerino_pivot_uscite[3],
#                          list_df_conti_camerino_pivot_uscite[4],
#                          list_df_conti_camerino_pivot_uscite[5],
#                          list_df_conti_camerino_pivot_uscite[6],
#                          list_df_conti_camerino_pivot_uscite[7],
#                          list_df_conti_camerino_pivot_uscite[8],
#                          list_df_conti_camerino_pivot_uscite[9],
#                          list_df_conti_camerino_pivot_uscite[10],
#                          list_df_conti_camerino_pivot_uscite[11]
#                          ]



    # Riportare  entrate del mese in ogni foglio
i = 0  # contatore list_mese
for sheet in list_ws_mese:
    for row in sheet.rows:
        for cell in row:
            if (cell.value == ("TOTALE_Uscite")):
                # print(cell.coordinate, cell.row, cell.column) # A184 184 1 per tutti e 12 i fogli
                # print(sheet.cell(row=cell.row, column=cell.column).value) # stampa 'TOTALE_Uscite
                sheet.cell(row=cell.offset(row=7, column=0).row, column=4,
                           value=list_df_conti_camerino_mese_entrate[i]['Euro'].sum(numeric_only=True))
                sheet.cell(row=cell.offset(row=7, column=0).row, column=4).font = Font(size=15, color='000000',
                                                                                       bold=True)
                sheet.cell(row=cell.offset(row=7, column=0).row, column=4).number_format = '#,##0.00€'
                sheet.cell(row=cell.offset(row=7, column=0).row, column=4).alignment = Alignment(horizontal="right")

    # Riportare  entrate del mese in ogni foglio

    for row in sheet.rows:
        for cell in row:
            if (cell.value == ("TOTALE_Uscite")):
                # print(cell.coordinate, cell.row, cell.column) # A184 184 1 per tutti e 12 i fogli
                # print(sheet.cell(row=cell.row, column=cell.column).value) # stampa 'TOTALE_Uscite
                sheet.cell(row=cell.offset(row=9, column=0).row, column=4,
                           value=list_df_conti_camerino_mese_uscite[i]['Euro'].sum(numeric_only=True))
                sheet.cell(row=cell.offset(row=9, column=0).row, column=4).font = Font(size=15, color='a81a1a',
                                                                                       bold=True)
                sheet.cell(row=cell.offset(row=9, column=0).row, column=4).number_format = '-#,##0.00€'
                sheet.cell(row=cell.offset(row=9, column=0).row, column=4).alignment = Alignment(horizontal="right")

    for row in sheet.rows:
        for cell in row:
            if (cell.value == ("TOTALE_Uscite")):
                # print(cell.coordinate, cell.row, cell.column) # A184 184 1 per tutti e 12 i fogli
                # print(sheet.cell(row=cell.row, column=cell.column).value) # stampa 'TOTALE_Uscite
                sheet.cell(row=cell.offset(row=11, column=0).row, column=4,
                           value=((list_df_conti_camerino_mese_entrate[i]['Euro']).sum(numeric_only=True) -
                                  (list_df_conti_camerino_mese_uscite[i]['Euro']).sum(numeric_only=True)
                                  )
                           )
                # sheet.cell(row=cell.offset(row=11, column=0).row, column=4).font = Font(size=15,
                #                                                                       bold=True)
                sheet.cell(row=cell.offset(row=11, column=0).row, column=4).number_format = '#,##0.00€'
                sheet.cell(row=cell.offset(row=11, column=0).row, column=4).alignment = Alignment(horizontal="right")

                if (sheet.cell(row=cell.offset(row=11, column=0).row, column=4).value) > 0:
                    sheet.cell(row=cell.offset(row=11, column=0).row, column=4).font = Font(color='000000', size=15,
                                                                                            bold=True)
                else:
                    sheet.cell(row=cell.offset(row=11, column=0).row, column=4).font = Font(color='a81a1a', size=15,
                                                                                            bold=True)

                i+= 1
##########

##################################################################################################################################

# saldo per tutti i mesi

#imposto saldo iniziale
saldo = 200_000
#imposto un contatore
e = 0

#saldo iniziale
for sheet in list_ws_mese:
    for row in sheet:
        for cell in row:
            if (cell.value == ("TOTALE_Uscite")):

                sheet.cell  (   row=cell.offset(row=5, column=0).row, column=4,
                                value=  (
                                                saldo
                                        )
                            )

                # mi calcolo il saldo finale e la assegno alla variabile saldo
                saldo = (saldo +
                         list_df_conti_camerino_mese_entrate[e]['Euro'].sum(numeric_only=True) -
                         list_df_conti_camerino_mese_uscite[e]['Euro'].sum(numeric_only=True)
                         )


                sheet.cell(row=cell.offset(row=5, column=0).row, column=4).font = Font(size=15, bold=True)
                sheet.cell(row=cell.offset(row=5, column=0).row, column=4).number_format = '#,##0.00€'
                sheet.cell(row=cell.offset(row=5, column=0).row, column=4).alignment = Alignment(horizontal="right")

                if sheet.cell(row=cell.offset(row=5, column=0).row, column=4).value > 0:
                        sheet.cell(row=cell.offset(row=5, column=0).row, column=4).font = Font(color='000000', size=15,
                                                                                        bold=True)
                else:
                        sheet.cell(row=cell.offset(row=5, column=0).row, column=4).font = Font(color='a81a1a', size=15,
                                                                                        bold=True)

                e += 1

# Saldo finale
    for row in sheet:
        for cell in row:
            if (cell.value == ("TOTALE_Uscite")):

                sheet.cell(row=cell.offset(row=13, column=0).row, column=4,
                            value=  (
                                        saldo
                                    )
                            )

                sheet.cell(row=cell.offset(row=13, column=0).row, column=4).font = Font(size=15, bold=True)
                sheet.cell(row=cell.offset(row=13, column=0).row, column=4).number_format = '#,##0.00€'
                sheet.cell(row=cell.offset(row=13, column=0).row, column=4).alignment = Alignment(horizontal="right")

                if sheet.cell(row=cell.offset(row=13, column=0).row, column=4).value > 0:
                    sheet.cell(row=cell.offset(row=13, column=0).row, column=4).font = Font(color='000000', size=15, bold=True)
                else:
                    sheet.cell(row=cell.offset(row=13, column=0).row, column=4).font = Font(color='a81a1a', size=15, bold=True)


# Salva
wb.save("conti_camerino_styled.xlsx")
